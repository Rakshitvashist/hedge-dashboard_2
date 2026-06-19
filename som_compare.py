"""
som_compare.py — Survivorship-bias comparison using the REAL SOM model.

Runs the Sharpe Single-Index Model (som_engine.run_som) on two universes:

  A) POINT-IN-TIME : each month only the stocks actually in Nifty 50 that month
                     are eligible (membership from nifty50_reshuffle_tracker_v2.xlsx),
                     priced from nifty50_host + nifty50_hist.
  B) CURRENT-FIXED : today's 50 Nifty stocks always eligible (nifty50_host) — the
                     survivorship-biased universe the live SOM uses.

Output: SOM_Survivorship_Comparison.xlsx, styled with SOM.py's colour palette.
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

from som_engine import run_som

# ── config ──
TRACKER     = "nifty50_reshuffle_tracker_v2.xlsx"
BENCH_FILE  = "NIFTY50_1d.csv"
BOND_FILE   = "India 1-Year Bond Yield Historical Data.csv"
# Match the production SOM window so Current-Fixed reproduces the live dashboard
# numbers (e.g. ~19% max drawdown). COVID-2020 is intentionally outside this window.
START_MONTH = "2021-04"
END_MONTH   = "2026-04"
ALIASES     = {"BAJAJAUTO": "BAJAJ-AUTO"}
OUT_FILE    = "SOM_Survivorship_Comparison.xlsx"


# ── point-in-time membership from the reshuffle tracker ──
def load_membership():
    fb = pd.read_excel(TRACKER, "Full Basket Per Month")
    months = [c for c in fb.columns if c != "Stock #"]
    mem = {}
    for m in months:
        names = fb[m].dropna().astype(str).str.strip().tolist()
        mem[m] = {ALIASES.get(n, n) for n in names}
    return mem


# ── SOM.py styling (verbatim palette + helpers) ──
def mkfill(h): return PatternFill("solid", fgColor=h)
def mkfont(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, name="Arial", italic=italic)
def mkborder(style='thin'):
    s = Side(style=style); return Border(top=s, bottom=s, left=s, right=s)

F_DARK = mkfill("1F3864"); F_MID = mkfill("2E75B6"); F_LIGHT = mkfill("D9E1F2")
F_GREEN = mkfill("E2EFDA"); F_DGRN = mkfill("C6EFCE"); F_RED = mkfill("FFE0E0")
F_DRED = mkfill("FFC7CE"); F_GRAY = mkfill("F2F2F2"); F_WHT = mkfill("FFFFFF")
F_TEAL = mkfill("003366"); F_MGOLD = mkfill("F4B942")
AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AR = Alignment(horizontal='right', vertical='center')
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)
FMT_PCT = '0.00%'; FMT_N2 = '0.00'; FMT_CASH = '#,##0'


def hdr(ws, r, c, val, bg=F_DARK, fg="FFFFFF", align=AC, size=9, bold=True, merge_to=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = mkfont(bold=bold, color=fg, size=size); cell.fill = bg
    cell.alignment = align; cell.border = mkborder()
    if merge_to: ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_to)
    return cell

def dat(ws, r, c, val, fill=F_WHT, fg="000000", align=AR, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = mkfont(bold=bold, color=fg); cell.fill = fill
    cell.alignment = align; cell.border = mkborder()
    if fmt: cell.number_format = fmt
    return cell


def yearly_returns(ms):
    """Compounded return per calendar year of trade_month."""
    df = ms.copy()
    df['year'] = df['trade_month'].str[:4]
    return df.groupby('year')['return_pct'].apply(lambda g: (1 + g).prod() - 1)


# ── build the styled workbook ──
def build_excel(pit, cur, coverage_pct):
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Survivorship Comparison")

    pm, cm = pit['metrics'], cur['metrics']

    # title
    ws.merge_cells('A1:E1')
    t = ws['A1']; t.value = "SHARPE SINGLE-INDEX MODEL  —  SURVIVORSHIP-BIAS COMPARISON"
    t.font = Font(bold=True, color="FFFFFF", size=14, name="Arial"); t.fill = F_DARK
    t.alignment = AC; ws.row_dimensions[1].height = 32
    ws.merge_cells('A2:E2')
    s = ws['A2']
    s.value = (f"Real SOM model run on both baskets  |  {pm['period']}  |  "
               f"{pm['months']} months  |  Point-in-time coverage {coverage_pct:.1f}%")
    s.font = Font(bold=True, color="375623", size=8, name="Arial"); s.fill = F_GREEN
    s.alignment = AC; s.border = mkborder()

    # metric table
    hdr(ws, 4, 1, "Metric", bg=F_TEAL, align=AL)
    hdr(ws, 4, 2, "Point-in-Time\n(real index)", bg=F_MID)
    hdr(ws, 4, 3, "Current-Fixed\n(survivorship)", bg=F_MID)
    hdr(ws, 4, 4, "Difference\n(PIT − Current)", bg=F_MGOLD, fg="000000")
    hdr(ws, 4, 5, "Benchmark\n(NIFTY 50)", bg=F_GRAY, fg="000000")

    rows = [
        ("Total Return",        'total_return',       FMT_PCT, 'bench_total_return'),
        ("CAGR (annualised)",   'ann_return',         FMT_PCT, 'bench_ann_return'),
        ("Annualised Volatility",'volatility',        FMT_PCT, None),
        ("Sharpe Ratio",        'sharpe',             FMT_N2,  None),
        ("Sortino Ratio",       'sortino',            FMT_N2,  None),
        ("Calmar Ratio",        'calmar',             FMT_N2,  None),
        ("Max Drawdown",        'max_drawdown',       FMT_PCT, None),
        ("Realized Beta",       'realized_beta',      FMT_N2,  None),
        ("Win Rate (months)",   'win_rate',           FMT_PCT, None),
        ("Avg # Stocks Held",   'avg_stocks',         FMT_N2,  None),
        ("Total Net PNL (Rs.)", 'total_net_pnl',      FMT_CASH, None),
        ("Final Capital (Rs.)", 'final_capital',      FMT_CASH, None),
    ]
    higher_is_worse = {'volatility', 'max_drawdown'}   # for diff colouring

    r = 5
    for label, key, fmt, bkey in rows:
        pv, cv = pm.get(key), cm.get(key)
        dat(ws, r, 1, label, fill=F_LIGHT, bold=True, align=AL)
        dat(ws, r, 2, pv, fmt=fmt)
        dat(ws, r, 3, cv, fmt=fmt)
        # difference (skip for ratios where it's less meaningful but still show)
        if isinstance(pv, (int, float)) and isinstance(cv, (int, float)) and not (pd.isna(pv) or pd.isna(cv)):
            diff = pv - cv
            good = diff > 0 if key not in higher_is_worse else diff < 0
            fill = F_DGRN if good else (F_DRED if diff != 0 else F_WHT)
            dat(ws, r, 4, diff, fmt=fmt, fill=fill, bold=True)
        else:
            dat(ws, r, 4, "—", align=AC)
        bval = pm.get(bkey) if bkey else None
        dat(ws, r, 5, bval if bval is not None else "—",
            fmt=fmt if bval is not None else None, fill=F_GRAY,
            align=AR if bval is not None else AC)
        r += 1

    # callout: survivorship bias
    r += 1
    bias = cm['ann_return'] - pm['ann_return']
    hdr(ws, r, 1, "SURVIVORSHIP BIAS (CAGR overstatement)", bg=F_TEAL, align=AL, merge_to=3)
    c = ws.cell(row=r, column=4, value=bias); c.font = mkfont(bold=True, size=11, color="9C0006")
    c.fill = F_DRED; c.alignment = AC; c.border = mkborder(); c.number_format = FMT_PCT
    dat(ws, r, 5, "per year", fill=F_GRAY, align=AC)

    for col, w in zip("ABCDE", [30, 16, 16, 16, 14]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[4].height = 30

    # ── yearly sheet ──
    wy = wb.create_sheet("Yearly Returns")
    wy.merge_cells('A1:D1')
    ty = wy['A1']; ty.value = "YEARLY RETURNS  —  Point-in-Time vs Current-Fixed"
    ty.font = Font(bold=True, color="FFFFFF", size=12, name="Arial"); ty.fill = F_DARK
    ty.alignment = AC; wy.row_dimensions[1].height = 26
    hdr(wy, 3, 1, "Year", bg=F_TEAL)
    hdr(wy, 3, 2, "Point-in-Time", bg=F_MID)
    hdr(wy, 3, 3, "Current-Fixed", bg=F_MID)
    hdr(wy, 3, 4, "Gap (PIT − Cur)", bg=F_MGOLD, fg="000000")
    py, cy = yearly_returns(pit['monthly_summary']), yearly_returns(cur['monthly_summary'])
    rr = 4
    for yr in sorted(set(py.index) | set(cy.index)):
        a, b = py.get(yr, np.nan), cy.get(yr, np.nan)
        dat(wy, rr, 1, yr, fill=F_LIGHT, bold=True, align=AC)
        dat(wy, rr, 2, a, fmt=FMT_PCT, fill=F_GREEN if a >= 0 else F_RED)
        dat(wy, rr, 3, b, fmt=FMT_PCT, fill=F_GREEN if b >= 0 else F_RED)
        gap = a - b
        dat(wy, rr, 4, gap, fmt=FMT_PCT, fill=F_DGRN if gap > 0 else F_DRED, bold=True)
        rr += 1
    for col, w in zip("ABCD", [10, 16, 16, 16]):
        wy.column_dimensions[col].width = w

    wb.save(OUT_FILE)
    return OUT_FILE


def main():
    membership = load_membership()

    print("Running SOM on CURRENT-FIXED universe (nifty50_host)...")
    cur = run_som(["nifty50_host"], BENCH_FILE, BOND_FILE, START_MONTH, END_MONTH,
                  membership=None, label="Current-Fixed")

    print("Running SOM on POINT-IN-TIME universe (nifty50_host + nifty50_hist)...")
    pit = run_som(["nifty50_host", "nifty50_hist"], BENCH_FILE, BOND_FILE,
                  START_MONTH, END_MONTH, membership=membership, label="Point-in-Time")

    # coverage = priceable share of each month's basket
    import os
    priceable = {f.replace("_1d_max.csv", "")
                 for d in ["nifty50_host", "nifty50_hist"] if os.path.isdir(d)
                 for f in os.listdir(d) if f.endswith("_1d_max.csv")}
    covs = []
    for m, basket in membership.items():
        if basket:
            covs.append(100 * len(basket & priceable) / len(basket))
    coverage_pct = float(np.mean(covs))

    out = build_excel(pit, cur, coverage_pct)

    pm, cm = pit['metrics'], cur['metrics']
    print("\n" + "=" * 60)
    print(f"  {'Metric':22} {'Point-in-Time':>16} {'Current-Fixed':>16}")
    print("-" * 60)
    for lab, k, f in [("CAGR", 'ann_return', '{:.2%}'), ("Total Return", 'total_return', '{:.2%}'),
                      ("Sharpe", 'sharpe', '{:.2f}'), ("Sortino", 'sortino', '{:.2f}'),
                      ("Max Drawdown", 'max_drawdown', '{:.2%}'), ("Avg # stocks", 'avg_stocks', '{:.1f}')]:
        print(f"  {lab:22} {f.format(pm[k]):>16} {f.format(cm[k]):>16}")
    print("-" * 60)
    print(f"  Survivorship bias (CAGR): {cm['ann_return']-pm['ann_return']:+.2%} per year")
    print(f"  Point-in-time coverage  : {coverage_pct:.1f}%")
    print(f"\n  Saved: {out}")


if __name__ == "__main__":
    main()
