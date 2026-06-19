"""
╔══════════════════════════════════════════════════════════════════════════╗
║   SHARPE SINGLE INDEX MODEL — UNIFIED BACKTESTER v1.0                    ║
║                                                                          ║
║   COMBINES:                                                              ║
║   • Script 1: Advanced PNL tracking with carry-forward logic             ║
║   • Script 1: Corrected beta calculation (intersection-based)            ║
║   • Script 1: Residual variance (theoretically accurate Sharpe model)    ║
║   • Script 2: Comprehensive Excel output with multiple dashboards        ║
║   • Script 2: Professional formatting, churning analysis, metrics        ║
║                                                                          ║
║   Features:                                                              ║
║   • Daily data → Monthly resampling                                      ║
║   • Zero look-ahead bias                                                 ║
║   • Realized PNL with continuity tracking                                ║
║   • Multiple summary dashboards (5Y, 3Y, 2Y, 1Y)                         ║
║   • Per-month portfolio sheets with correlation matrices                 ║
║   • Churning analysis (stocks added/removed)                             ║
║   • Full metrics: Sharpe, Sortino, Calmar, Max Drawdown                  ║
║                                                                          ║
║                                                                          ║
╚══════════════════════════════════════════════════════════════════════════╝
"""

import os, sys, warnings
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

warnings.filterwarnings("ignore")


# ==============================================================================
# [Config] CONFIGURATION (Script 2 style - easy to modify)
# ==============================================================================

# Input folders and files
# Point-in-time (reshuffle) universe: current basket prices + fetched historical
# constituents. Each month only the stocks ACTUALLY in Nifty 500 that month are
# eligible (membership from the reshuffle tracker) -> no survivorship bias.
STOCKS_FOLDERS   = ["nifty500_host", "nifty500_hist"]  # loaded together
STOCKS_FOLDER    = STOCKS_FOLDERS[0]                    # kept for any legacy refs
MEMBERSHIP_FILE  = "nifty500_reshuffle_tracker_v2.xlsx"
MEMBERSHIP_ALIAS = {"BAJAJAUTO": "BAJAJ-AUTO"}
BOND_FILE        = "India 1-Year Bond Yield Historical Data.csv"  # Risk-free rate
BENCHMARK_FILE   = "NIFTY500_1d.csv"  # Benchmark index

# Output files
OUTPUT_FILE      = "SOM_Reshuffle500_PointInTime.xlsx"          # Main comprehensive report
SUMMARY_ONLY_FILE = "SOM_Reshuffle500_PointInTime_Summary.xlsx" # Summary-only report
CURRENT_PORT_FILE = "SOM_Reshuffle500_PointInTime_Current.xlsx" # Latest trading instructions

# Backtest period (matches the hosted SOM window)
START_MONTH      = "2021-04"
END_MONTH        = "2026-04"

# Portfolio parameters
BETA_WINDOW      = 60           # Months for beta calculation (adaptive up to this max)
MIN_OBS          = 2          # Minimum observations required (variance needs at least 2)
MAX_WEIGHT       = 0.10         # Maximum weight per stock (10%)
INVESTMENT       = 10_000_000   # Total investment amount (Rs. 1 Crore)
TXN_COST_RATE    = 0.002        # Transaction cost (0.2%)

# Technical parameters
EPSILON          = 1e-10        # Small number for division safety


# ══════════════════════════════════════════════════════════════════════════
# 🛠  HELPER FUNCTIONS (Combined from both scripts)
# ══════════════════════════════════════════════════════════════════════════

def read_daily_csv(file_path):
    """
    Read daily CSV file with robust date parsing (Script 1 + Script 2 combined)
    Handles multiple date formats, column name variations, and comma-formatted numbers
    """
    try:
        df = pd.read_csv(file_path)
    except Exception:
        df = pd.read_csv(file_path, sep='\t')
    
    # Normalize column names: strip whitespace + lowercase
    df.columns = df.columns.str.strip().str.lower()
    
    # Rename date column variants to 'time'
    # Expanded list to handle various data sources (Investing.com, NSE, Yahoo, etc.)
    date_col_variants = ['date', 'time', 'timestamp', 'datetime', 'trade_date', 'trading_date', 'day', 'period']
    for variant in date_col_variants:
        if variant in df.columns:
            df.rename(columns={variant: 'time'}, inplace=True)
            break
    
    # Rename price column variants to 'close'
    # Investing.com bond/index exports use 'price' instead of 'close'
    if 'price' in df.columns and 'close' not in df.columns:
        df.rename(columns={'price': 'close'}, inplace=True)
    
    if 'time' not in df.columns:
        raise ValueError(f"No date column found in {file_path}")
    
    # Flexible date parsing
    orig_time = df['time'].copy()
    # First try YYYY-MM-DD
    df['time'] = pd.to_datetime(orig_time, format='%Y-%m-%d', errors='coerce')
    
    # Handle remaining dates based on separator (Slashed in this dataset is M/D/YYYY, Hyphenated is DD-MM-YYYY)
    mask = df['time'].isna()
    if mask.any():
        sample = orig_time[mask].astype(str).head(10)
        has_hyphen = sample.str.contains('-').any()
        if has_hyphen:
            df.loc[mask, 'time'] = pd.to_datetime(orig_time[mask], dayfirst=True, errors='coerce')
        else:
            df.loc[mask, 'time'] = pd.to_datetime(orig_time[mask], dayfirst=False, errors='coerce')
    
    # Remove commas from numeric columns (Script 2 addition)
    numeric_cols = ['open', 'high', 'low', 'close', 'volume']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = (
                df[col].astype(str)
                .str.replace(',', '', regex=False)
                .pipe(pd.to_numeric, errors='coerce')
            )
    
    # Clean and sort
    df = df.dropna(subset=['time', 'close'])
    df = df.sort_values('time').reset_index(drop=True)
    
    return df


def resample_to_monthly(daily_df):
    """
    Convert daily OHLC data to monthly (Script 1 approach)
    First trading day's open, month's high/low, last trading day's close
    """
    df = daily_df.set_index('time')
    
    # Aggregation rules
    agg_dict = {
        'open': 'first',   # First day's open price
        'high': 'max',     # Highest price in the month
        'low': 'min',      # Lowest price in the month
        'close': 'last'    # Last day's close price
    }
    
    # Only aggregate columns that exist
    agg_dict = {k: v for k, v in agg_dict.items() if k in df.columns}
    
    # Resample to month-end
    monthly = df.resample('ME').agg(agg_dict).reset_index()
    
    # Add year_month identifier
    monthly['year_month'] = monthly['time'].dt.to_period('M')
    
    # Calculate monthly returns
    monthly['return'] = monthly['close'].pct_change().fillna(0)
    monthly['Monthly Returns'] = monthly['return'] * 100  # Percentage
    
    return monthly.set_index('year_month').sort_index()


def calculate_avg_price(previous_avg_price, previous_qty, current_qty, open_price):
    """
    Calculate weighted average price when adding positions (Script 1)
    """
    if current_qty == 0:
        return 0
    
    previous_value = previous_avg_price * previous_qty
    added_qty = abs(current_qty - previous_qty)
    added_value = added_qty * open_price
    
    return (previous_value + added_value) / current_qty


def cap_weights_iterative(raw_weights, cap, max_iter=1000):
    """
    Iteratively cap weights and redistribute excess (Script 1 + Script 2 combined)
    """
    w = np.array(raw_weights, dtype=float).copy()
    
    for _ in range(max_iter):
        over = w > cap
        if not over.any():
            break
        
        excess = (w[over] - cap).sum()
        w[over] = cap
        under = ~over
        
        if not under.any():
            break
        
        under_sum = w[under].sum()
        if under_sum > EPSILON:
            w[under] += excess * (w[under] / under_sum)
    
    # Normalize to sum=1
    total = w.sum()
    if total > EPSILON:
        w = w / total
    
    return w


# ══════════════════════════════════════════════════════════════════════════
# 📂  LOAD AND PROCESS DATA
# ══════════════════════════════════════════════════════════════════════════

print("=" * 80)
print("  SHARPE UNIFIED BACKTESTER v1.0 -- LOADING DATA")
print("  Combining Script 1's calculation logic + Script 2's output format")
print("=" * 80)

# Load benchmark
print(f"\n[*] Loading benchmark: {BENCHMARK_FILE}")
bench_daily = read_daily_csv(BENCHMARK_FILE)
bench_monthly = resample_to_monthly(bench_daily)
bench_daily_idx = bench_daily.set_index('time').sort_index()
bench_daily_idx['return'] = bench_daily_idx['close'].pct_change()
print(f"    [v] Benchmark loaded: {len(bench_monthly)} monthly bars")

# Load bond (risk-free rate)
print(f"[$$] Loading risk-free rate: {BOND_FILE}")
bond_daily = read_daily_csv(BOND_FILE)
bond_monthly = resample_to_monthly(bond_daily)
print(f"    [v] Bond data loaded: {len(bond_monthly)} monthly bars")

# Load point-in-time membership (which stocks were in Nifty 50 each month)
print(f"\n[dir] Loading membership: {MEMBERSHIP_FILE}")
_fb = pd.read_excel(MEMBERSHIP_FILE, "Full Basket Per Month")
MEMBERSHIP = {}
for _m in [c for c in _fb.columns if c != "Stock #"]:
    _names = _fb[_m].dropna().astype(str).str.strip().tolist()
    MEMBERSHIP[_m] = {MEMBERSHIP_ALIAS.get(n, n) for n in _names}
print(f"    [v] Membership loaded: {len(MEMBERSHIP)} months")

# Load stocks from all universe folders (clean ticker = filename minus _1d_max)
print(f"\n[dir] Loading stocks from: {STOCKS_FOLDERS}")
all_stocks = {}
all_stocks_daily = {}
load_errors = []

for folder in STOCKS_FOLDERS:
    if not os.path.isdir(folder):
        print(f"  [!] folder not found: {folder}")
        continue
    for file in sorted(f for f in os.listdir(folder) if f.lower().endswith('.csv')):
        name = Path(file).stem.replace('_1d_max', '').replace('_1d', '').strip()
        if name in all_stocks:        # first folder wins (current prices preferred)
            continue
        try:
            daily = read_daily_csv(os.path.join(folder, file))
            daily = daily.drop_duplicates(subset='time', keep='last')  # guard reindex
            monthly = resample_to_monthly(daily)
            all_stocks[name] = monthly
            all_stocks_daily[name] = daily.set_index('time')
        except Exception as e:
            load_errors.append((name, str(e)))
            print(f"  [x]  {name:25s}  ERROR: {e}")

if not all_stocks:
    sys.exit("❌  No stock CSVs loaded from any folder")

print(f"\n[v] Successfully loaded {len(all_stocks)} stocks")
if load_errors:
    print(f"[!] {len(load_errors)} files failed to load")


# ==============================================================================
# [Loop]  PORTFOLIO GENERATION LOOP (Script 1's superior calculation logic)
# ==============================================================================

print(f"\n{'='*80}")
print(f"  WALK-FORWARD PORTFOLIO GENERATION")
print(f"  Period: {START_MONTH} -- {END_MONTH}")
print(f"  Investment: Rs.{INVESTMENT:,.0f}")
print(f"  Max Weight: {MAX_WEIGHT:.0%} per stock")
print(f"{'='*80}\n")

all_port_months = pd.period_range(start=START_MONTH, end=END_MONTH, freq='M')
monthly_portfolios = []
daily_log = []
skipped_months = []

# Port state for carry-forward tracking (Script 1)
port_state = {}  # symbol -> {'qty': int, 'avg_price': float, 'last_close': float}

# Churning records (Script 2)
churn_records = []

# Full EGP tables and correlation matrices (Script 2)
full_egp_tables = {}
full_correlation_matrices = {}

for port_month in all_port_months:
    trade_month = port_month + 1
    
    # Check if risk-free rate available
    if port_month not in bond_monthly.index:
        skipped_months.append((str(port_month), "RF rate missing"))
        continue
    
    risk_free_rate = float(bond_monthly.loc[port_month, 'close']) / 100.0
    
    # Get benchmark history
    bench_hist = bench_monthly[bench_monthly.index <= port_month]
    bench_window = bench_hist.tail(BETA_WINDOW)
    
    if len(bench_window) < MIN_OBS:
        skipped_months.append((str(port_month), f"Benchmark only {len(bench_window)} months"))
        continue
    
    # Calculate benchmark variance (Script 1: uses Monthly Returns column)
    bench_var_annual = bench_window['return'].var(ddof=1) * 12
    
    if pd.isna(bench_var_annual) or bench_var_annual < EPSILON:
        skipped_months.append((str(port_month), "Benchmark variance too low or NaN"))
        continue
    
    # ──────────────────────────────────────────────────────────────────────
    # Calculate beta and variance for each stock (Script 1's approach)
    # ──────────────────────────────────────────────────────────────────────
    
    records = []

    # Point-in-time eligibility: only stocks that were actually in Nifty 50 during
    # the month we will HOLD (trade_month) may be selected. This removes survivorship bias.
    eligible = MEMBERSHIP.get(str(trade_month))

    for symbol, stock_monthly in all_stocks.items():
        if eligible is not None and symbol not in eligible:
            continue
        # Get historical data up to portfolio month
        stock_hist = stock_monthly[stock_monthly.index <= port_month]
        
        # Try monthly check first
        stock_window = stock_hist.tail(min(BETA_WINDOW, len(stock_hist)))
        common_idx = stock_window.index.intersection(bench_window.index)
        
        use_daily = len(common_idx) < MIN_OBS
        
        if use_daily:
            # ✅ DAILY FALLBACK for newly listed stocks
            s_daily = all_stocks_daily.get(symbol, pd.DataFrame())
            if s_daily.empty: continue
            
            # Filter up to portfolio month
            end_date = port_month.to_timestamp(how='end')
            s_daily_hist = s_daily[s_daily.index <= end_date]
            
            if len(s_daily_hist) < 5: continue # Minimum 5 days for beta
            
            s_returns = s_daily_hist['close'].pct_change().dropna()
            m_returns = bench_daily_idx['return'].reindex(s_returns.index).dropna()
            
            common_idx_daily = s_returns.index.intersection(m_returns.index)
            if len(common_idx_daily) < 5: continue
            
            stock_returns = s_returns.loc[common_idx_daily].values
            market_returns = m_returns.loc[common_idx_daily].values
            ann_factor = 252 # Annualize from daily
            obs_count = len(common_idx_daily)
        else:
            stock_returns = stock_window.loc[common_idx, 'return'].values
            market_returns = bench_window.loc[common_idx, 'return'].values
            ann_factor = 12 # Annualize from monthly
            obs_count = len(common_idx)
            
        # Calculate covariance matrix
        cov_matrix = np.cov(stock_returns, market_returns, ddof=1)
        covariance = cov_matrix[0, 1]
        market_variance = cov_matrix[1, 1]
        
        if market_variance < EPSILON:
            continue
            
        # Calculate beta
        beta = covariance / market_variance
        
        if pd.isna(beta) or beta <= 0:
            continue
        
        # ✅ RESIDUAL VARIANCE (Theoretically correct Sharpe model)
        # Calculate alpha (intercept)
        alpha = stock_returns.mean() - beta * market_returns.mean()
        
        # Calculate residuals (unsystematic component)
        residuals = stock_returns - (alpha + beta * market_returns)
        
        # Residual variance (annualized)
        stock_variance = np.var(residuals, ddof=1) * ann_factor
        
        if pd.isna(stock_variance) or stock_variance < EPSILON:
            continue
        
        # Expected return (annualized)
        annual_ret = stock_returns.mean() * ann_factor
        
        # ──────────────────────────────────────────────────────────────────
        # Get trading prices (Handle Current Portfolio Selection)
        # ──────────────────────────────────────────────────────────────────
        
        if trade_month in stock_monthly.index:
            trade_row = stock_monthly.loc[trade_month]
            buy_price = trade_row.get('open', np.nan)
            sell_price = trade_row.get('close', np.nan)
            
            # If buy_price is NaN but month exists, use last close
            if pd.isna(buy_price) or buy_price <= 0:
                buy_price = stock_hist['close'].iloc[-1]
            if pd.isna(sell_price):
                sell_price = buy_price
            trade_data_exists = True
        else:
            # Future Selection Mode: Use last known close as buy price for qty calculation
            buy_price = stock_hist['close'].iloc[-1]
            sell_price = buy_price
            trade_data_exists = False
        
        if pd.isna(buy_price) or buy_price <= 0:
            continue
        
        # Store data
        records.append({
            'symbol': symbol,
            'beta': beta,
            'annual_ret': annual_ret,
            'sigma2': stock_variance,
            'erb': (annual_ret - risk_free_rate) / beta,
            'n_obs': obs_count,
            'buy_price': buy_price,
            'sell_price': sell_price,
            'trade_data_exists': trade_data_exists
        })
    
    if not records:
        skipped_months.append((str(port_month), "No qualified stocks"))
        continue
    
    # ──────────────────────────────────────────────────────────────────────
    # Calculate Sharpe Single Index Model (Script 1 + Script 2 combined)
    # ──────────────────────────────────────────────────────────────────────
    
    df_s = pd.DataFrame(records).sort_values('erb', ascending=False).reset_index(drop=True)
    
    # Calculate Sharpe factors
    df_s['A'] = ((df_s['annual_ret'] - risk_free_rate) * df_s['beta']) / df_s['sigma2']
    df_s['H'] = df_s['beta'] ** 2 / df_s['sigma2']
    df_s['cum_A'] = df_s['A'].cumsum()
    df_s['cum_H'] = df_s['H'].cumsum()
    
    # Calculate cutoff rate (Ci)
    denom = 1 + bench_var_annual * df_s['cum_H']
    if (denom.abs() < EPSILON).any():
        skipped_months.append((str(port_month), "Near-zero Ci denominator"))
        continue
    
    df_s['Ci'] = (bench_var_annual * df_s['cum_A']) / denom
    
    # Find cutoff point
    if df_s['Ci'].isna().all():
        skipped_months.append((str(port_month), "All Ci values are NaN"))
        continue
        
    cutoff_pos = int(df_s['Ci'].idxmax())
    C_star = float(df_s.loc[cutoff_pos, 'Ci'])
    n_selected = cutoff_pos + 1
    
    # Save full EGP table
    full_egp_tables[str(port_month)] = df_s.copy()
    
    # Select stocks above cutoff
    sel = df_s.iloc[:n_selected].copy()
    
    # Calculate weights
    sel['Zi'] = ((sel['beta'] / sel['sigma2']) * (sel['erb'] - C_star)).clip(lower=0)
    Z_sum = sel['Zi'].sum()
    
    if Z_sum < EPSILON:
        skipped_months.append((str(port_month), "Z_sum ≈ 0"))
        continue
    
    sel['wi_raw'] = sel['Zi'] / Z_sum
    sel['wi'] = cap_weights_iterative(sel['wi_raw'].values, MAX_WEIGHT)
    
    # Calculate quantities
    sel['allocation'] = INVESTMENT * sel['wi']
    sel['qty'] = np.floor(sel['allocation'] / sel['buy_price']).astype(int)
    sel = sel[sel['qty'] > 0].copy()
    
    if sel.empty:
        skipped_months.append((str(port_month), "All qty = 0"))
        continue
    
    # ──────────────────────────────────────────────────────────────────────
    # Ex-ante Portfolio Statistics (Daily Data for better coverage)
    # ──────────────────────────────────────────────────────────────────────
    
    selected_symbols = sel['symbol'].tolist()
    weights = sel['wi'].values
    returns_daily_list = []
    
    # We use daily data for ex-ante risk to handle new stocks with < 1 month history
    end_date_limit = port_month.to_timestamp(how='end')
    
    for sym in selected_symbols:
        if sym in all_stocks_daily:
            s_daily = all_stocks_daily[sym]
            s_daily_hist = s_daily[s_daily.index <= end_date_limit].tail(BETA_WINDOW * 21)
            s_ret = s_daily_hist['close'].pct_change().rename(sym)
            returns_daily_list.append(s_ret)
    
    if returns_daily_list:
        # Align daily returns
        ret_matrix = pd.concat(returns_daily_list, axis=1).dropna()
        
        if not ret_matrix.empty and len(ret_matrix) >= 5: # Need at least 5 common days
            cov_mat_annual = ret_matrix.cov() * 252 # Annualize daily covariance
            port_variance = weights.T @ cov_mat_annual.values @ weights
            ex_ante_vol = np.sqrt(max(0, port_variance))
            port_exp_ret = (sel['annual_ret'] * sel['wi']).sum()
            ex_ante_sr = (port_exp_ret - risk_free_rate) / ex_ante_vol if ex_ante_vol > EPSILON else 0.0
            ex_ante_corr = ret_matrix.corr()
        else:
            ex_ante_vol, ex_ante_sr = 0.0, 0.0
            ex_ante_corr = pd.DataFrame()
    else:
        ex_ante_vol, ex_ante_sr = 0.0, 0.0
        ex_ante_corr = pd.DataFrame()
    
    sel['ex_ante_vol'] = ex_ante_vol
    sel['ex_ante_sharpe'] = ex_ante_sr
    full_correlation_matrices[str(port_month)] = ex_ante_corr
    
    # Portfolio Beta: Weighted average of individual stock betas
    # Formula: Beta_p = Sum(w_i * Beta_i)
    port_beta = (sel['beta'] * sel['wi']).sum()
    sel['port_beta'] = port_beta
    
    # ──────────────────────────────────────────────────────────────────────
    # Churning Analysis (Script 2) + Carry-Forward Setup (Script 1)
    # ──────────────────────────────────────────────────────────────────────
    
    target_stocks = set(sel['symbol'].tolist())
    prev_stocks = set(port_state.keys())
    
    added_list = sorted(list(target_stocks - prev_stocks))
    removed_list = sorted(list(prev_stocks - target_stocks))
    remained_list = sorted(list(target_stocks & prev_stocks))
    
    churn_records.append({
        'port_month': str(port_month),
        'trade_month': str(trade_month),
        'added_count': len(added_list),
        'removed_count': len(removed_list),
        'remained_count': len(remained_list),
        'added_symbols': ", ".join(added_list),
        'removed_symbols': ", ".join(removed_list),
        'remained_symbols': ", ".join(remained_list),
        'total_stocks': len(target_stocks)
    })
    
    sel['status'] = sel['symbol'].apply(lambda x: 'Added' if x in added_list else 'Remained')
    
    # ──────────────────────────────────────────────────────────────────────
    # Calculate PNL with Carry-Forward Logic (Script 1)
    # ──────────────────────────────────────────────────────────────────────
    
    sel['prev_qty'] = sel['symbol'].map(lambda x: port_state.get(x, {}).get('qty', 0))
    sel['prev_avg_price'] = sel['symbol'].map(lambda x: port_state.get(x, {}).get('avg_price', 0.0))
    sel['prev_last_close'] = sel['symbol'].map(lambda x: port_state.get(x, {}).get('last_close', 0.0))
    sel['delta_qty'] = sel['qty'] - sel['prev_qty']
    
    # Action needed (Trading instructions) - MUST BE AFTER delta_qty
    def get_action(row):
        dq = row['delta_qty']
        if dq > 0: return f"BUY {int(dq)}"
        elif dq < 0: return f"SELL {int(abs(dq))}"
        else: return "HOLD"
    
    sel['action_needed'] = sel.apply(get_action, axis=1)
    
    # Corrected Average Price (True Carry-Forward)
    sel['avg_price'] = sel.apply(
        lambda row: calculate_avg_price(
            row['prev_avg_price'], row['prev_qty'], row['qty'], row['buy_price']
        ) if row['qty'] > row['prev_qty'] else (row['prev_avg_price'] if row['prev_qty'] > 0 else row['buy_price']),
        axis=1
    )
    
    # PNL Components
    # 1. Gap PNL: Change from last month's close to this month's open (for shares held into the month)
    # If it's a new stock, gap is 0
    sel.loc[sel['prev_qty'] == 0, 'prev_last_close'] = sel.loc[sel['prev_qty'] == 0, 'buy_price']
    sel['gap_pnl'] = (sel['buy_price'] - sel['prev_last_close']) * sel['prev_qty']
    
    # 2. Current Month PNL: Change from this month's open to close (for shares currently held)
    sel['current_pnl'] = (sel['sell_price'] - sel['buy_price']) * sel['qty']
    
    # 3. Realized PNL from entirely removed stocks
    current_removed_pnl = 0.0
    exit_value_removed = 0.0 # for txn cost
    for sym in removed_list:
        if sym in port_state:
            if trade_month in all_stocks[sym].index:
                # Removed at the OPEN of trade month
                exit_price = all_stocks[sym].loc[trade_month, 'open']
                p_close = port_state[sym]['last_close']
                p_avg = port_state[sym]['avg_price']
                p_qty = port_state[sym]['qty']
                current_removed_pnl += (exit_price - p_close) * p_qty
                exit_value_removed += p_qty * exit_price
            else:
                # Future Selection Mode: No PNL yet
                pass
    
    sel['removed_pnl'] = current_removed_pnl
    sel['gross_pnl'] = sel['gap_pnl'] + sel['current_pnl']
    
    # Investment values for reporting
    sel['invest'] = sel['avg_price'] * sel['qty']
    sel['exit_val'] = sel['qty'] * sel['sell_price']
    
    # Transaction costs
    # Entry cost (for new positions or additions)
    entry_value = (sel['delta_qty'].clip(lower=0) * sel['buy_price']).sum()
    
    # Exit cost (for positions being removed)
    exit_value = exit_value_removed
    
    # Reduction cost (for positions being reduced - uses buy_price as it happens at start of month)
    reduction_value = (sel['delta_qty'].clip(upper=0).abs() * sel['buy_price']).sum()
    
    total_txn_cost = (entry_value + exit_value + reduction_value) * TXN_COST_RATE
    
    # Distribute cost proportionally
    sel['cost'] = total_txn_cost * (sel['invest'] / sel['invest'].sum()) if sel['invest'].sum() > 0 else 0
    
    # Net PNL (Including removed stocks)
    monthly_net_pnl = sel['gross_pnl'].sum() + current_removed_pnl - total_txn_cost
    
    # If this is a future selection (no trade data), zero out the PNL metrics
    if not sel['trade_data_exists'].iloc[0]:
        monthly_net_pnl = 0.0
        sel['gross_pnl'] = 0.0
        sel['gap_pnl'] = 0.0
        sel['current_pnl'] = 0.0
        sel['removed_pnl'] = 0.0
        sel['cost'] = 0.0
        total_txn_cost = 0.0
        current_removed_pnl = 0.0
    
    # Distribute cost and removed PNL proportionally for stock-level reporting
    sel['net_pnl'] = sel['gross_pnl'] - sel['cost']
    sel['return_pct'] = sel['net_pnl'] / sel['invest'] if sel['invest'].sum() > 0 else 0.0
    
    # Update port_state for next month
    for _, row in sel.iterrows():
        port_state[row['symbol']] = {
            'qty': row['qty'],
            'avg_price': row['avg_price'],
            'last_close': row['sell_price'],
            'buy_price': row['buy_price'],
            'prev_last_close': row['prev_last_close'],
            'prev_qty': row['prev_qty']
        }
    
    # Remove exited positions
    for sym in removed_list:
        port_state.pop(sym, None)
    
    # Add metadata
    sel['port_month'] = str(port_month)
    sel['trade_month'] = str(trade_month)
    sel['rf_rate'] = risk_free_rate
    sel['C_star'] = C_star
    sel['bench_var'] = bench_var_annual
    sel['added_count'] = len(added_list)
    sel['removed_count'] = len(removed_list)
    sel['added_symbols'] = ", ".join(added_list)
    sel['removed_symbols'] = ", ".join(removed_list)
    
    # Benchmark return for the trade month
    if trade_month in bench_monthly.index:
        sel['bench_trade_ret'] = bench_monthly.loc[trade_month, 'return']
    else:
        sel['bench_trade_ret'] = 0.0
    
    total_gap_pnl = sel['gap_pnl'].sum()
    
    monthly_portfolios.append(sel)

    # --- DAILY EQUITY TRACKING ---
    trade_month_str = str(trade_month)
    month_mask = (bench_daily['time'].dt.strftime('%Y-%m') == trade_month_str)
    month_dates = bench_daily[month_mask]['time'].sort_values()
    
    # Pre-fetch daily prices for all current symbols to avoid repeated indexing
    daily_prices_cache = {}
    for sym in port_state.keys():
        if sym in all_stocks_daily:
            # Reindex to month_dates and ffill to handle missing data gracefully
            s_data = all_stocks_daily[sym]['close']
            # Intersection to avoid errors if some dates are completely missing
            available_dates = s_data.index.intersection(month_dates)
            if not available_dates.empty:
                daily_prices_cache[sym] = s_data.loc[available_dates].reindex(month_dates).ffill()
    
    for day in month_dates:
        m2m_pnl_this_month = 0.0
        for sym, state in port_state.items():
            qty = state['qty']
            buy_price = state['buy_price']
            
            # Default to buy_price if no daily data yet
            day_close = buy_price
            if sym in daily_prices_cache and day in daily_prices_cache[sym].index:
                price = daily_prices_cache[sym].loc[day]
                if not pd.isna(price):
                    day_close = price
            
            m2m_pnl_this_month += (day_close - buy_price) * qty
            
        # daily_net_pnl is the gain WITHIN this month
        daily_net_pnl = m2m_pnl_this_month + total_gap_pnl + current_removed_pnl - total_txn_cost
        daily_wealth_this_month = INVESTMENT + daily_net_pnl
        
        bench_mask = (bench_daily['time'] == day)
        b_close = bench_daily.loc[bench_mask, 'close'].iloc[0] if bench_mask.any() else 1.0
        
        daily_log.append({
            'date': day,
            'trade_month': trade_month_str,
            'wealth_this_month': daily_wealth_this_month,
            'bench_close': b_close
        })

    # Print progress
    ti = sel['invest'].sum()
    tp = monthly_net_pnl
    print(f"  [{port_month}] -- {trade_month}  "
          f"stocks={n_selected:3d}  C*={C_star:+.4f}  RF={risk_free_rate:.2%}  "
          f"invested=Rs.{ti/1e5:.1f}L  net_pnl={'+'if tp>=0 else ''}Rs.{tp/1e5:.1f}L")

if skipped_months:
    print(f"\n  Skipped {len(skipped_months)} month(s):")
    for m, r in skipped_months:
        print(f"    [x]  {m}  -- {r}")

if not monthly_portfolios:
    sys.exit("[x]  No valid portfolios generated!")


# ==============================================================================
# [Metrics]  CALCULATE SUMMARIES AND METRICS
# ==============================================================================

print(f"\n{'='*80}")
print("  CALCULATING PERFORMANCE METRICS")
print(f"{'='*80}")

combined = pd.concat(monthly_portfolios, ignore_index=True)

monthly_summary = (
    combined.groupby('port_month')
    .agg(
        trade_month=('trade_month', 'first'),
        n_stocks=('symbol', 'count'),
        rf_rate=('rf_rate', 'first'),
        bench_ret=('bench_trade_ret', 'first'),
        invest=('invest', 'sum'),
        exit_val=('exit_val', 'sum'),
        gross_pnl=('gross_pnl', 'sum'),
        removed_pnl=('removed_pnl', 'first'),
        cost=('cost', 'sum'),
        ex_ante_vol=('ex_ante_vol', 'first'),
        ex_ante_sr=('ex_ante_sharpe', 'first'),
        port_beta=('port_beta', 'first'),
        added_count=('added_count', 'first'),
        removed_count=('removed_count', 'first'),
        added_symbols=('added_symbols', 'first'),
        removed_symbols=('removed_symbols', 'first'),
    ).reset_index()
)

monthly_summary['net_pnl'] = monthly_summary['gross_pnl'] + monthly_summary['removed_pnl'] - monthly_summary['cost']
monthly_summary['return_pct'] = monthly_summary['net_pnl'] / INVESTMENT
monthly_summary['cumulative_pnl'] = monthly_summary['net_pnl'].cumsum()
monthly_summary['cumulative_ret'] = (1 + monthly_summary['return_pct']).cumprod() - 1
monthly_summary['bench_cum_ret'] = (1 + monthly_summary['bench_ret']).cumprod() - 1
monthly_summary['alpha'] = monthly_summary['return_pct'] - monthly_summary['bench_ret']
monthly_summary['beat_bench'] = monthly_summary['alpha'] > 0
monthly_summary['compounded_cap'] = INVESTMENT * (1 + monthly_summary['cumulative_ret'])

# Overall metrics
total_months = len(monthly_summary)

# Realized Portfolio Beta (Overall)
# Formula: Beta = Cov(Rp, Rm) / Var(Rm)
if total_months > 1:
    cov_matrix = np.cov(monthly_summary['return_pct'], monthly_summary['bench_ret'], ddof=1)
    realized_beta_overall = cov_matrix[0, 1] / cov_matrix[1, 1] if cov_matrix[1, 1] > EPSILON else 0.0
else:
    realized_beta_overall = 0.0
win_months = (monthly_summary['net_pnl'] > 0).sum()
total_net_pnl = monthly_summary['net_pnl'].sum()
ann_return = (1 + monthly_summary['cumulative_ret'].iloc[-1]) ** (12 / total_months) - 1
volatility = monthly_summary['return_pct'].std() * (12 ** 0.5)
rf_avg = monthly_summary['rf_rate'].mean()

# Sharpe Ratio
monthly_summary['excess_return'] = monthly_summary['return_pct'] - (monthly_summary['rf_rate'] / 12)
excess_ann_return = monthly_summary['excess_return'].mean() * 12
sharpe_ratio = excess_ann_return / volatility if volatility > EPSILON else 0.0

# Max Drawdown
wealth_series = pd.concat([pd.Series([1.0]), 1 + monthly_summary['cumulative_ret']])
peak = wealth_series.cummax()
drawdown = (wealth_series - peak) / peak
max_drawdown = drawdown.min()

# Sortino Ratio
downside_returns = monthly_summary['excess_return'][monthly_summary['excess_return'] < 0]
if len(downside_returns) > 0:
    downside_dev = np.sqrt(np.mean(downside_returns ** 2)) * np.sqrt(12)
    sortino_ratio = excess_ann_return / downside_dev if downside_dev > EPSILON else np.nan
else:
    sortino_ratio = np.nan

# Calmar Ratio
calmar_ratio = ann_return / abs(max_drawdown) if abs(max_drawdown) > EPSILON else np.nan

# Churning DataFrame
churn_df = pd.DataFrame(churn_records)

print(f"[v] Metrics calculated for {total_months} months")


# ==============================================================================
# [Style]  EXCEL FORMATTING HELPERS (Script 2's comprehensive styling)
# ==============================================================================

def mkfill(h): return PatternFill("solid", fgColor=h)
def mkfont(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, name="Arial", italic=italic)
def mkborder(style='thin'):
    s = Side(style=style)
    return Border(top=s, bottom=s, left=s, right=s)

# Color palette
F_DARK = mkfill("1F3864");  F_MID = mkfill("2E75B6");  F_LIGHT = mkfill("D9E1F2")
F_GREEN = mkfill("E2EFDA");  F_DGRN = mkfill("C6EFCE");  F_RED = mkfill("FFE0E0")
F_DRED = mkfill("FFC7CE");  F_YLW = mkfill("FFFF00");  F_ORG = mkfill("FCE4D6")
F_GRAY = mkfill("F2F2F2");  F_WHT = mkfill("FFFFFF");  F_WARN = mkfill("FFF3CD")
F_TEAL = mkfill("003366");  F_LBLUE = mkfill("BDD7EE");  F_MGOLD = mkfill("F4B942")

# Alignments
AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AR = Alignment(horizontal='right', vertical='center')
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Number formats
FMT_PCT = '0.00%';  FMT_N4 = '0.0000';  FMT_N6 = '0.000000'
FMT_CASH = '#,##0';  FMT_C2 = '#,##0.00';  FMT_INT = '0'

def hdr(ws, row, col, val, bold=True, bg=F_DARK, fg="FFFFFF",
        align=AC, size=9, fmt=None, merge_to=None):
    """Create a header cell"""
    c = ws.cell(row=row, column=col, value=val)
    c.font = mkfont(bold=bold, color=fg, size=size)
    c.fill = bg
    c.alignment = align
    c.border = mkborder()
    if fmt: c.number_format = fmt
    if merge_to: ws.merge_cells(start_row=row, start_column=col,
                                 end_row=row, end_column=merge_to)
    return c

def dat(ws, row, col, val, fill=F_WHT, fg="000000",
        align=AR, fmt=None, bold=False, merge_to=None):
    """Create a data cell"""
    c = ws.cell(row=row, column=col, value=val)
    c.font = mkfont(bold=bold, color=fg)
    c.fill = fill
    c.alignment = align
    c.border = mkborder()
    if fmt: c.number_format = fmt
    if merge_to: ws.merge_cells(start_row=row, start_column=col,
                                 end_row=row, end_column=merge_to)
    return c


# ==============================================================================
# [Report]  BUILD EXCEL OUTPUT (Script 2's comprehensive format)
# ==============================================================================

print(f"\n{'='*80}")
print("  BUILDING EXCEL OUTPUT")
print(f"{'='*80}\n")

wb = Workbook()
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

wb_summary = Workbook()
if "Sheet" in wb_summary.sheetnames:
    del wb_summary["Sheet"]


def build_summary_dashboard(ws, sub, title, stats, churn_rows,
                           sub_total_months, sub_ann_return, sub_bench_ann_return):
    """Build a summary dashboard sheet"""
    
    # Title
    ws.merge_cells('A1:T1')
    ws['A1'].value = f"SHARPE UNIFIED MODEL — {title.upper()} PERFORMANCE"
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
    ws['A1'].fill = F_DARK
    ws['A1'].alignment = AC
    ws.row_dimensions[1].height = 34
    
    # Subtitle
    ws.merge_cells('A2:T2')
    ws['A2'].value = (
        f"Period: Last {sub_total_months} months  |  Zero Look-ahead Bias  |  "
        "BUY at OPEN of trade month  |  SELL at CLOSE of trade month"
    )
    ws['A2'].font = Font(bold=True, color="375623", size=8, name="Arial")
    ws['A2'].fill = F_GREEN
    ws['A2'].alignment = AC
    ws['A2'].border = mkborder()
    
    # Stats boxes
    for j, (label, val) in enumerate(stats, 1):
        hdr(ws, 4, j, label, bg=F_MID, size=8)
        c = ws.cell(row=5, column=j, value=val)
        c.font = mkfont(bold=True, size=10)
        c.fill = F_LIGHT
        c.alignment = AC
        c.border = mkborder()
        ws.column_dimensions[gcl(j)].width = 16
    
    # Churning stats
    hdr(ws, 4, 11, "Churning Stats (Stocks)", merge_to=14, bg=F_TEAL)
    hdr(ws, 5, 11, "Metric", bg=F_MID, size=8)
    hdr(ws, 5, 12, "Max", bg=F_MID, size=8)
    hdr(ws, 5, 13, "Min", bg=F_MID, size=8)
    hdr(ws, 5, 14, "Avg", bg=F_MID, size=8)
    
    p_mx, p_mn, p_av = sub['n_stocks'].max(), sub['n_stocks'].min(), sub['n_stocks'].mean()
    expanded_churn = churn_rows + [("Portfolio Size", p_mx, p_mn, p_av)]
    
    for i, (lab, mx, mn, av) in enumerate(expanded_churn, 6):
        dat(ws, i, 11, lab, fill=F_LIGHT, bold=True, align=AL)
        dat(ws, i, 12, mx, fill=F_WHT, fmt=FMT_INT)
        dat(ws, i, 13, mn, fill=F_WHT, fmt=FMT_INT)
        dat(ws, i, 14, av, fill=F_WHT, fmt=FMT_C2)
    
    # Top churned symbols
    hdr(ws, 4, 16, "Most Frequent Churns", merge_to=18, bg=F_MGOLD)
    hdr(ws, 5, 16, "Status", bg=F_MID, size=8)
    hdr(ws, 5, 17, "Top Symbols (freq)", merge_to=18, bg=F_MID, size=8)
    
    all_add, all_rem = [], []
    for s in sub['added_symbols']:
        if s: all_add.extend([x.strip() for x in s.split(",")])
    for s in sub['removed_symbols']:
        if s: all_rem.extend([x.strip() for x in s.split(",")])
    
    def get_top_str(l):
        if not l: return "N/A"
        vc = pd.Series(l).value_counts().head(3)
        return ", ".join([f"{k}({v})" for k, v in vc.items()])
    
    dat(ws, 6, 16, "Top Added", fill=F_LIGHT, bold=True, align=AL)
    dat(ws, 6, 17, get_top_str(all_add), merge_to=18, align=AL)
    dat(ws, 7, 16, "Top Removed", fill=F_LIGHT, bold=True, align=AL)
    dat(ws, 7, 17, get_top_str(all_rem), merge_to=18, align=AL)
    
    # Main table headers
    sum_headers = [
        "Portfolio\nMonth", "Trade\nMonth", "Stocks", "Added\nStocks", "Removed\nStocks", "RF Rate",
        "Ex-ante\nBeta", "Invest (Rs.)", "Exit Value (Rs.)", "Net PnL (Rs.)", "Port\nReturn %",
        "Bench\nReturn %", "Alpha %", "Beat\nBench?", "Ex-ante\nVol", "Ex-ante\nSharpe",
        "Port\nCumul %", "Bench\nCumul %", "Compounded\nCapital (Rs.)",
        "Added Symbols", "Removed Symbols"
    ]
    header_widths = [13, 13, 7, 8, 8, 8, 10, 14, 14, 14, 11, 11, 10, 10, 10, 10, 12, 12, 16, 40, 40]
    
    for j, h in enumerate(sum_headers, 1):
        hdr(ws, 10, j, h)
        ws.column_dimensions[gcl(j)].width = header_widths[j-1]
    ws.row_dimensions[10].height = 42
    
    # Data rows
    for i, row in sub.iterrows():
        r = i + 11
        pnl = row['net_pnl']
        fill = F_GREEN if pnl >= 0 else F_RED
        fg = "375623" if pnl >= 0 else "9C0006"
        
        dat(ws, r, 1, row['port_month'], fill=fill, align=AC)
        dat(ws, r, 2, row['trade_month'], fill=fill, align=AC)
        dat(ws, r, 3, int(row['n_stocks']), fill=fill, align=AC, fmt=FMT_INT)
        dat(ws, r, 4, int(row['added_count']), fill=fill, align=AC, fmt=FMT_INT)
        dat(ws, r, 5, int(row['removed_count']), fill=fill, align=AC, fmt=FMT_INT)
        dat(ws, r, 6, row['rf_rate'], fill=fill, fmt=FMT_PCT)
        dat(ws, r, 7, row['port_beta'], fill=F_GRAY, fmt=FMT_N4)
        dat(ws, r, 8, row['invest'], fill=fill, fmt=FMT_CASH)
        dat(ws, r, 9, row['exit_val'], fill=fill, fmt=FMT_CASH)
        dat(ws, r, 10, pnl, fill=fill, fmt=FMT_CASH, bold=True, fg=fg)
        dat(ws, r, 11, row['return_pct'], fill=fill, fmt=FMT_PCT)
        
        dat(ws, r, 12, row['bench_ret'], fill=F_GRAY, fmt=FMT_PCT)
        alpha_fill = F_DGRN if row['alpha'] > 0 else F_DRED
        dat(ws, r, 13, row['alpha'], fill=alpha_fill, fmt=FMT_PCT)
        dat(ws, r, 14, "YES" if row['beat_bench'] else "NO", fill=alpha_fill, align=AC)
        
        dat(ws, r, 15, row['ex_ante_vol'], fill=F_GRAY, fmt=FMT_PCT)
        dat(ws, r, 16, row['ex_ante_sr'], fill=F_GRAY, fmt=FMT_N4)
        
        dat(ws, r, 17, row['sub_cum_ret'], fill=fill, fmt=FMT_PCT)
        dat(ws, r, 18, row['sub_bench_cum_ret'], fill=F_GRAY, fmt=FMT_PCT)
        dat(ws, r, 19, row['sub_comp_cap'], fill=fill, fmt=FMT_CASH, bold=True)
        
        dat(ws, r, 20, row['added_symbols'], fill=fill, align=AL)
        dat(ws, r, 21, row['removed_symbols'], fill=fill, align=AL)
    
    # Growth comparison table
    hdr(ws, 10, 22, "Growth of Rs.100", merge_to=24)
    hdr(ws, 11, 22, "Month", bg=F_MID)
    hdr(ws, 11, 23, "Portfolio", bg=F_MID)
    hdr(ws, 11, 24, "Benchmark", bg=F_MID)
    ws.column_dimensions[gcl(22)].width = 16
    ws.column_dimensions[gcl(23)].width = 14
    ws.column_dimensions[gcl(24)].width = 14
    
    for i, row in sub.iterrows():
        r = i + 12
        dat(ws, r, 22, row['trade_month'], align=AC)
        dat(ws, r, 23, 100 * (1 + row['sub_cum_ret']), fmt=FMT_C2)
        dat(ws, r, 24, 100 * (1 + row['sub_bench_cum_ret']), fmt=FMT_C2)
    
    # Totals row
    tr = len(sub) + 12
    hdr(ws, tr, 1, "TOTAL", bg=F_DARK, merge_to=5)
    for col, val, fmt in [
        (8, sub['invest'].sum(), FMT_CASH),
        (9, sub['exit_val'].sum(), FMT_CASH),
        (10, sub['net_pnl'].sum(), FMT_CASH),
        (11, sub_ann_return, FMT_PCT),
        (12, sub_bench_ann_return, FMT_PCT),
        (13, sub_ann_return - sub_bench_ann_return, FMT_PCT),
        (7, sub['port_beta'].mean(), FMT_N4),
        (15, sub['ex_ante_vol'].mean(), FMT_PCT),
        (16, sub['ex_ante_sr'].mean(), FMT_N4),
        (17, sub['sub_cum_ret'].iloc[-1], FMT_PCT),
        (18, sub['sub_bench_cum_ret'].iloc[-1], FMT_PCT),
        (19, sub['sub_comp_cap'].iloc[-1], FMT_CASH),
    ]:
        c = ws.cell(row=tr, column=col, value=val)
        c.font = mkfont(bold=True, color="FFFF00", size=9)
        c.fill = F_DARK
        c.alignment = AR
        c.border = mkborder()
        c.number_format = fmt
    
    ws.freeze_panes = 'A11'


# Create summary dashboards
summary_configs = [
    ("Summary 5Y", 60),
    ("Summary 3Y", 36),
    ("Summary 2Y", 24),
    ("Summary 1Y", 12),
]

for title, lookback in summary_configs:
    sub = monthly_summary.tail(lookback).copy().reset_index(drop=True)
    if sub.empty:
        continue
    
    sub_total_months = len(sub)
    sub_win_months = (sub['net_pnl'] > 0).sum()
    
    sub['sub_cum_ret'] = (1 + sub['return_pct']).cumprod() - 1
    sub['sub_bench_cum_ret'] = (1 + sub['bench_ret']).cumprod() - 1
    sub['sub_comp_cap'] = INVESTMENT * (1 + sub['sub_cum_ret'])
    
    sub_ann_return = (1 + sub['sub_cum_ret'].iloc[-1]) ** (12 / sub_total_months) - 1
    sub_bench_ann_return = (1 + sub['sub_bench_cum_ret'].iloc[-1]) ** (12 / sub_total_months) - 1
    
    wealth_sub = pd.concat([pd.Series([1.0]), 1 + sub['sub_cum_ret']])
    peak_sub = wealth_sub.cummax()
    sub_max_drawdown = ((wealth_sub - peak_sub) / peak_sub).min()
    
    sub_avg_ex_sr = sub['ex_ante_sr'].mean()
    sub_stability = (sub['added_count'] == 0).sum() / sub_total_months
    
    s_max_add, s_min_add, s_avg_add = sub['added_count'].max(), sub['added_count'].min(), sub['added_count'].mean()
    s_max_rem, s_min_rem, s_avg_rem = sub['removed_count'].max(), sub['removed_count'].min(), sub['removed_count'].mean()
    s_remained = sub['n_stocks'] - sub['added_count']
    s_max_con, s_min_con, s_avg_con = s_remained.max(), s_remained.min(), s_remained.mean()
    
    # Realized Portfolio Beta for sub-period
    if sub_total_months > 1:
        cov_matrix_sub = np.cov(sub['return_pct'], sub['bench_ret'], ddof=1)
        sub_realized_beta = cov_matrix_sub[0, 1] / cov_matrix_sub[1, 1] if cov_matrix_sub[1, 1] > EPSILON else 0.0
    else:
        sub_realized_beta = 0.0

    stats = [
        ("Months Traded", f"{sub_total_months}"),
        ("Win Rate (Pnl)", f"{sub_win_months/sub_total_months:.1%}"),
        ("Selection Stability", f"{sub_stability:.1%}"),
        ("Port Ann. Ret", f"{sub_ann_return:.2%}"),
        ("Bench Ann. Ret", f"{sub_bench_ann_return:.2%}"),
        ("Avg Ex-ante Beta", f"{sub['port_beta'].mean():.4f}"),
        ("Realized Port Beta", f"{sub_realized_beta:.4f}"),
        ("Avg Ex-ante Sharpe", f"{sub_avg_ex_sr:.3f}"),
        ("Max Drawdown", f"{sub_max_drawdown:.2%}"),
        ("Total Net PnL", f"Rs.{sub['net_pnl'].sum():,.0f}"),
    ]
    
    churn_rows = [
        ("Added", s_max_add, s_min_add, s_avg_add),
        ("Removed", s_max_rem, s_min_rem, s_avg_rem),
        ("Remained", s_max_con, s_min_con, s_avg_con),
    ]
    
    ws_main = wb.create_sheet(title)
    build_summary_dashboard(ws_main, sub, title, stats, churn_rows,
                           sub_total_months, sub_ann_return, sub_bench_ann_return)
    
    ws_only = wb_summary.create_sheet(title)
    build_summary_dashboard(ws_only, sub, title, stats, churn_rows,
                           sub_total_months, sub_ann_return, sub_bench_ann_return)
    
    print(f"  [v] Created summary: {title}")


# ══════════════════════════════════════════════════════════════════════════════
# CHURNING ANALYSIS SHEET
# ══════════════════════════════════════════════════════════════════════════════

ws_churn = wb.create_sheet("Churning Analysis")

ws_churn.merge_cells('A1:I1')
ws_churn['A1'].value = "PORTFOLIO CHURNING ANALYSIS (SYMBOL CHANGES)"
ws_churn['A1'].font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
ws_churn['A1'].fill = F_DARK
ws_churn['A1'].alignment = AC
ws_churn.row_dimensions[1].height = 26

# Summary statistics
hdr(ws_churn, 3, 1, "CHURNING SUMMARY STATISTICS", merge_to=5, bg=F_TEAL)
hdr(ws_churn, 4, 1, "Metric", bg=F_MID, size=8)
hdr(ws_churn, 4, 2, "Max (Value)", bg=F_MID, size=8)
hdr(ws_churn, 4, 3, "Max (Month)", bg=F_MID, size=8)
hdr(ws_churn, 4, 4, "Min (Month)", bg=F_MID, size=8)
hdr(ws_churn, 4, 5, "Average", bg=F_MID, size=8)

for i, (label, col_name) in enumerate([("Added Stocks", "added_count"),
                                      ("Removed Stocks", "removed_count"),
                                      ("Remained Stocks", "remained_count")], 5):
    max_val = churn_df[col_name].max()
    min_val = churn_df[col_name].min()
    avg_val = churn_df[col_name].mean()
    max_month = churn_df.loc[churn_df[col_name].idxmax(), 'port_month']
    min_month = churn_df.loc[churn_df[col_name].idxmin(), 'port_month']
    
    dat(ws_churn, i, 1, label, fill=F_LIGHT, bold=True, align=AL)
    dat(ws_churn, i, 2, max_val, align=AC, fmt=FMT_INT)
    dat(ws_churn, i, 3, max_month, align=AC)
    dat(ws_churn, i, 4, min_month, align=AC)
    dat(ws_churn, i, 5, avg_val, align=AC, fmt=FMT_C2)

# Main churning table
c_hdrs = ["Port\nMonth", "Trade\nMonth", "Added\nCount", "Added Symbols",
          "Removed\nCount", "Removed Symbols", "Remained\nCount", "Remained Symbols", "Total\nStocks"]
c_wids = [13, 12, 8, 40, 8, 40, 8, 40, 8]

tbl_start_row = 10
for j, h in enumerate(c_hdrs, 1):
    hdr(ws_churn, tbl_start_row, j, h, bg=F_MID)
    ws_churn.column_dimensions[gcl(j)].width = c_wids[j-1]
ws_churn.row_dimensions[tbl_start_row].height = 40

for i, row in churn_df.iterrows():
    r = i + tbl_start_row + 1
    dat(ws_churn, r, 1, row['port_month'], align=AC)
    dat(ws_churn, r, 2, row['trade_month'], align=AC)
    dat(ws_churn, r, 3, row['added_count'], align=AC, fmt=FMT_INT)
    dat(ws_churn, r, 4, row['added_symbols'], align=AL)
    dat(ws_churn, r, 5, row['removed_count'], align=AC, fmt=FMT_INT)
    dat(ws_churn, r, 6, row['removed_symbols'], align=AL)
    dat(ws_churn, r, 7, row['remained_count'], align=AC, fmt=FMT_INT)
    dat(ws_churn, r, 8, row['remained_symbols'], align=AL)
    dat(ws_churn, r, 9, row['total_stocks'], align=AC, fmt=FMT_INT)

ws_churn.freeze_panes = f'A{tbl_start_row + 1}'
print(f"  [v] Created sheet: Churning Analysis")


# ══════════════════════════════════════════════════════════════════════════════
# STOCK DETAIL SHEET
# ══════════════════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Stock Detail")

ws2.merge_cells('A1:U1')
ws2['A1'].value = "STOCK-LEVEL DETAIL -- ALL MONTHS"
ws2['A1'].font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
ws2['A1'].fill = F_DARK
ws2['A1'].alignment = AC
ws2.row_dimensions[1].height = 26

d_hdrs = ["Port\nMonth", "Trade\nMonth", "Symbol", "Status", "N Obs",
          "Beta", "Ann Return", "Sigma²", "ERB", "C*", "Zi", "Wi\n(raw)", "Wi\n(capped)",
          "RF Rate", "Avg Price", "Buy Price", "Sell Price", "Qty",
          "Invested (Rs.)", "Exit Val (Rs.)", "Net PnL (Rs.)", "Return %"]
d_cols = ['port_month', 'trade_month', 'symbol', 'status', 'n_obs',
          'beta', 'annual_ret', 'sigma2', 'erb', 'C_star', 'Zi', 'wi_raw', 'wi',
          'rf_rate', 'avg_price', 'buy_price', 'sell_price', 'qty',
          'invest', 'exit_val', 'net_pnl', 'return_pct']
d_fmts = [None, None, None, None, FMT_INT,
          FMT_N4, FMT_PCT, FMT_N6, FMT_N4, FMT_N6, FMT_N6, FMT_PCT, FMT_PCT,
          FMT_PCT, FMT_C2, FMT_C2, FMT_C2, FMT_INT,
          FMT_CASH, FMT_CASH, FMT_CASH, FMT_PCT]
d_wids = [13, 12, 14, 10, 6, 9, 10, 11, 9, 10, 10, 8, 9, 8, 11, 11, 11, 8, 14, 14, 14, 9]

for j, h in enumerate(d_hdrs, 1):
    hdr(ws2, 2, j, h)
    ws2.column_dimensions[gcl(j)].width = d_wids[j-1]
ws2.row_dimensions[2].height = 40

for i, (_, row) in enumerate(combined.iterrows()):
    r = i + 3
    pnl = row['net_pnl']
    fill = F_GREEN if pnl >= 0 else F_RED
    fg = "375623" if pnl >= 0 else "9C0006"
    for j, (col, fmt) in enumerate(zip(d_cols, d_fmts), 1):
        val = int(row[col]) if col == 'qty' else row[col]
        bold = col in ('net_pnl',)
        dat(ws2, r, j, val, fill=fill, fmt=fmt,
            align=AC if j <= 4 else AR, bold=bold, fg=fg if bold else "000000")

ws2.freeze_panes = 'A3'
print(f"  [v] Created sheet: Stock Detail")


# ══════════════════════════════════════════════════════════════════════════════
# EGP AUDIT SHEET
# ══════════════════════════════════════════════════════════════════════════════

ws3 = wb.create_sheet("EGP Audit")

ws3.merge_cells('A1:N1')
ws3['A1'].value = "EGP ALGORITHM AUDIT — FULL TABLE PER MONTH"
ws3['A1'].font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
ws3['A1'].fill = F_DARK
ws3['A1'].alignment = AC
ws3.row_dimensions[1].height = 24

e_hdrs = ["Port Month", "Symbol", "Rank", "Beta", "Ann Ret",
          "Sigma²", "ERB", "A", "Cum A", "H", "Cum H", "Ci", "C*", "Selected?"]
e_wids = [13, 16, 6, 9, 10, 11, 9, 13, 13, 13, 13, 13, 10, 10]
e_fmts = [None, None, FMT_INT, FMT_N4, FMT_PCT, FMT_N6, FMT_N4,
          FMT_N6, FMT_N6, FMT_N6, FMT_N6, FMT_N6, FMT_N6, None]

for j, h in enumerate(e_hdrs, 1):
    hdr(ws3, 2, j, h)
    ws3.column_dimensions[gcl(j)].width = e_wids[j-1]
ws3.row_dimensions[2].height = 36

egp_row = 3
for pm_str, df_egp in full_egp_tables.items():
    c_star = df_egp['Ci'].max()
    
    for i, row in df_egp.iterrows():
        is_cs = abs(row['Ci'] - c_star) < 1e-8
        fill = F_YLW if is_cs else F_LIGHT
        vals = [pm_str, row['symbol'], i+1, row['beta'], row['annual_ret'],
                row['sigma2'], row['erb'], row['A'], row['cum_A'], row['H'],
                row['cum_H'], row['Ci'], c_star, "YES [DONE]" if is_cs else ""]
        for j, (v, fmt) in enumerate(zip(vals, e_fmts), 1):
            c = ws3.cell(row=egp_row, column=j, value=v)
            c.font = mkfont(bold=is_cs)
            c.fill = fill
            c.alignment = AC if j <= 3 else AR
            c.border = mkborder()
            if fmt: c.number_format = fmt
        egp_row += 1

ws3.freeze_panes = 'A3'
print(f"  [v] Created sheet: EGP Audit")


# ══════════════════════════════════════════════════════════════════════════════
# BIAS AUDIT SHEET
# ══════════════════════════════════════════════════════════════════════════════

ws4 = wb.create_sheet("Bias Audit")

ws4.merge_cells('A1:E1')
ws4['A1'].value = "BIAS & CHEATING AUDIT REPORT"
ws4['A1'].font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
ws4['A1'].fill = F_DARK
ws4['A1'].alignment = AC
ws4.row_dimensions[1].height = 30

for j, (h, w) in enumerate(zip(
    ["#", "Check", "Description", "Status", "How It's Handled"],
    [4, 22, 52, 14, 60]), 1):
    hdr(ws4, 2, j, h)
    ws4.column_dimensions[gcl(j)].width = w
ws4.row_dimensions[2].height = 22

audit_items = [
    ("Look-ahead in returns", "Stock/benchmark returns use future data?", "PASS [DONE]",
     "stock_hist filtered: index <= port_month before calculating returns"),
    ("Look-ahead in RF rate", "Risk-free rate from a future period?", "PASS [DONE]",
     f"RF = bond_monthly.loc[port_month, 'close'] / 100 (same month M)"),
    ("Look-ahead in beta", "Beta window extends past port_month?", "PASS [DONE]",
     f"Adaptive window: min({BETA_WINDOW}, available data) with index <= port_month"),
    ("Look-ahead in bench var", "σ²_m uses future benchmark data?", "PASS [DONE]",
     f"bench_var_annual calculated from bench_window (all data <= port_month)"),
    ("Trade price look-ahead", "Future prices used to decide portfolio weights?", "PASS [DONE]",
     "Buy/sell from trade_month M+1 — used only for execution, NOT for signal"),
    ("Weight capping", "Iterative capping may leave residual over-allocation", "PASS [DONE]",
     f"Iterative loop up to 1000 iterations; breaks only when no wi > {MAX_WEIGHT:.0%}"),
    ("Transaction costs", "Costs modelled realistically?", "PASS [DONE]",
     f"{TXN_COST_RATE:.1%} on entries/exits/adjustments"),
    ("Negative beta filter", "Inverse-market stocks excluded?", "PASS [DONE]",
     "beta <= 0 stocks skipped; long-only universe"),
    ("Minimum observations", "Beta estimated from too few data points?", "PASS [DONE]",
     f"Minimum {MIN_OBS} aligned monthly observations required"),
    ("Resampling correctness", "Monthly open/close from correct daily prices?", "PASS [DONE]",
     "resample('ME'): open='first' (first day), close='last' (last day)"),
    ("Residual variance", "Uses theoretically correct Sharpe model?", "PASS [DONE]",
     "Residual variance calculated from (return - alpha - beta*market_return)"),
    ("Carry-forward PNL", "PNL tracking handles continuity correctly?", "PASS [DONE]",
     "Weighted average price for additions, last close for continuations"),
    ("Intersection-based beta", "Beta calculated on aligned data only?", "PASS [DONE]",
     "Common months identified via index.intersection() before covariance"),
    ("Survivorship bias", "Delisted/failed stocks excluded?", "WARN ⚠️",
     "Add delisted/merged stocks to STOCKS_FOLDER for unbiased results"),
]

for i, (check, desc, status, handling) in enumerate(audit_items, 3):
    is_pass = "PASS" in status
    is_warn = "WARN" in status
    fill = F_DGRN if is_pass else (F_WARN if is_warn else F_DRED)
    sfg = "375623" if is_pass else ("7D6608" if is_warn else "9C0006")
    dat(ws4, i, 1, i-2, fill=fill, align=AC, fmt=FMT_INT)
    dat(ws4, i, 2, check, fill=fill, align=AL)
    dat(ws4, i, 3, desc, fill=fill, align=AL)
    dat(ws4, i, 4, status, fill=fill, align=AC, fg=sfg, bold=True)
    dat(ws4, i, 5, handling, fill=fill, align=AL)
    ws4.row_dimensions[i].height = 24

ws4.freeze_panes = 'A3'
print(f"  [v] Created sheet: Bias Audit")


# ══════════════════════════════════════════════════════════════════════════════
# MONTH INDEX SHEET (Clickable Navigator)
# ══════════════════════════════════════════════════════════════════════════════

ws_idx = wb.create_sheet("📋 Month Index")

ws_idx.merge_cells('A1:G1')
ws_idx['A1'].value = "MONTHLY PORTFOLIO INDEX — Click a month to see its portfolio sheet"
ws_idx['A1'].font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
ws_idx['A1'].fill = F_TEAL
ws_idx['A1'].alignment = AC
ws_idx.row_dimensions[1].height = 32

idx_hdrs = ["Sheet Name", "Port Month", "Trade Month", "# Stocks",
            "RF Rate", "Net PnL (Rs.)", "Monthly Return"]
idx_wids = [18, 14, 14, 10, 10, 16, 14]

for j, (h, w) in enumerate(zip(idx_hdrs, idx_wids), 1):
    hdr(ws_idx, 2, j, h, bg=F_MID)
    ws_idx.column_dimensions[gcl(j)].width = w
ws_idx.row_dimensions[2].height = 22

for i, pm_df in enumerate(monthly_portfolios):
    r = i + 3
    pm_str = pm_df['port_month'].iloc[0]
    tm_str = pm_df['trade_month'].iloc[0]
    sheet_nm = f"PM_{pm_str}"
    n_stocks = len(pm_df)
    rf_r = pm_df['rf_rate'].iloc[0]
    net_pnl = pm_df['net_pnl'].sum()
    ret_pct = net_pnl / pm_df['invest'].sum()
    fill = F_GREEN if net_pnl >= 0 else F_RED
    fg = "375623" if net_pnl >= 0 else "9C0006"
    
    # Hyperlink to the per-month sheet
    link_cell = ws_idx.cell(row=r, column=1, value=sheet_nm)
    link_cell.hyperlink = f"#{sheet_nm}!A1"
    link_cell.font = Font(color="0563C1", underline="single",
                          bold=True, size=9, name="Arial")
    link_cell.fill = fill
    link_cell.alignment = AC
    link_cell.border = mkborder()
    
    dat(ws_idx, r, 2, pm_str, fill=fill, align=AC)
    dat(ws_idx, r, 3, tm_str, fill=fill, align=AC)
    dat(ws_idx, r, 4, n_stocks, fill=fill, align=AC, fmt=FMT_INT)
    dat(ws_idx, r, 5, rf_r, fill=fill, fmt=FMT_PCT)
    dat(ws_idx, r, 6, net_pnl, fill=fill, fmt=FMT_CASH, bold=True, fg=fg)
    dat(ws_idx, r, 7, ret_pct, fill=fill, fmt=FMT_PCT)

ws_idx.freeze_panes = 'A3'
print(f"  [v] Created sheet: Month Index")


# ══════════════════════════════════════════════════════════════════════════════
# PER-MONTH PORTFOLIO SHEETS (PM_YYYY-MM)
# ══════════════════════════════════════════════════════════════════════════════

print(f"\n[doc] Creating {len(monthly_portfolios)} per-month portfolio sheets...")

PM_HDRS = [
    "Rank", "Symbol", "Status", "N Obs", "Beta", "Ann Ret %",
    "Sigma²", "ERB", "Zi", "Wi Raw %", "Wi Capped %",
    "Avg Price\n(Rs.)", "Buy Price\n(Rs.)", "Sell Price\n(Rs.)", "Qty",
    "Invested\n(Rs.)", "Exit Val\n(Rs.)", "Gross PnL\n(Rs.)",
    "Txn Cost\n(Rs.)", "Net PnL\n(Rs.)", "Return %"
]
PM_COLS = [
    None, 'symbol', 'status', 'n_obs', 'beta', 'annual_ret',
    'sigma2', 'erb', 'Zi', 'wi_raw', 'wi',
    'avg_price', 'buy_price', 'sell_price', 'qty',
    'invest', 'exit_val', 'gross_pnl',
    'cost', 'net_pnl', 'return_pct'
]
PM_FMTS = [
    FMT_INT, None, None, FMT_INT, FMT_N4, FMT_PCT,
    FMT_N6, FMT_N4, FMT_N6, FMT_PCT, FMT_PCT,
    FMT_C2, FMT_C2, FMT_C2, FMT_INT,
    FMT_CASH, FMT_CASH, FMT_CASH,
    FMT_CASH, FMT_CASH, FMT_PCT
]
PM_WIDS = [
    6, 16, 10, 6, 9, 10,
    11, 9, 10, 9, 10,
    12, 12, 12, 8,
    14, 14, 14,
    12, 14, 10
]
NCOLS = len(PM_HDRS)

for pm_df in monthly_portfolios:
    pm_str = pm_df['port_month'].iloc[0]
    tm_str = pm_df['trade_month'].iloc[0]
    rf_r = pm_df['rf_rate'].iloc[0]
    c_star = pm_df['C_star'].iloc[0]
    bv = pm_df['bench_var'].iloc[0]
    bt_ret = pm_df['bench_trade_ret'].iloc[0]
    
    sheet_nm = f"PM_{pm_str}"
    ws = wb.create_sheet(sheet_nm)
    
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    c = ws.cell(row=1, column=1,
                value=f"PORTFOLIO  ·  {pm_str}  |  Data used: all months ≤ {pm_str}  |  "
                      f"Trades execute in: {tm_str}")
    c.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    c.fill = F_TEAL
    c.alignment = AC
    c.border = mkborder()
    ws.row_dimensions[1].height = 30
    
    # Anti-cheat banner
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    c = ws.cell(row=2, column=1,
                value=(f"[DONE]  ZERO LOOK-AHEAD  |  EGP weights built using data ≤ {pm_str} only  |  "
                       f"BUY price = OPEN of 1st trading day of {tm_str}  |  "
                       f"SELL price = CLOSE of last trading day of {tm_str}"))
    c.font = Font(bold=True, color="375623", size=8, name="Arial")
    c.fill = F_GREEN
    c.alignment = AC
    c.border = mkborder()
    ws.row_dimensions[2].height = 16
    
    # Parameter summary block
    params = [
        ("Portfolio Month", pm_str), ("Trade Month", tm_str),
        ("RF Rate (annual)", f"{rf_r:.4%}"), ("C* (cutoff)", f"{c_star:.6f}"),
        ("Bench Var (annual)", f"{bv:.6f}"), ("Bench Ret (month)", f"{bt_ret:.2%}"),
        ("Ex-ante Vol", f"{pm_df['ex_ante_vol'].iloc[0]:.2%}"),
        ("Ex-ante Sharpe", f"{pm_df['ex_ante_sharpe'].iloc[0]:.3f}"),
        ("Portfolio Beta", f"{pm_df['port_beta'].iloc[0]:.4f}"),
        ("Investment Budget", f"Rs.{INVESTMENT:,.0f}"),
        ("Max Weight Cap", f"{MAX_WEIGHT:.0%}"),
    ]
    
    param_row = 4
    for idx, (label, value) in enumerate(params):
        col_offset = (idx % 4) * 2 + 1
        if idx % 4 == 0 and idx > 0:
            param_row += 2
        hdr(ws, param_row, col_offset, label, bg=F_MID, size=8, fg="FFFFFF",
            merge_to=col_offset)
        dat(ws, param_row+1, col_offset, value, fill=F_LBLUE, bold=True, align=AC)
        ws.column_dimensions[gcl(col_offset)].width = PM_WIDS[min(idx*2, NCOLS-1)]
    
    # Spacer label
    tbl_start = 11
    ws.merge_cells(start_row=tbl_start, start_column=1,
                   end_row=tbl_start, end_column=NCOLS)
    c = ws.cell(row=tbl_start, column=1,
                value=f"[v]  SELECTED PORTFOLIO STOCKS  ({len(pm_df)} stocks)")
    c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    c.fill = F_DARK
    c.alignment = AL
    c.border = mkborder()
    ws.row_dimensions[tbl_start].height = 22
    
    # Table headers
    tbl_hdr_row = tbl_start + 1
    for j, (h, w) in enumerate(zip(PM_HDRS, PM_WIDS), 1):
        hdr(ws, tbl_hdr_row, j, h, bg=F_MID)
        ws.column_dimensions[gcl(j)].width = w
    ws.row_dimensions[tbl_hdr_row].height = 42
    
    # Data rows
    pm_sorted = pm_df.sort_values('wi', ascending=False).reset_index(drop=True)
    for i, row in pm_sorted.iterrows():
        r = tbl_hdr_row + 1 + i
        pnl = row['net_pnl']
        fill = F_GREEN if pnl >= 0 else F_RED
        fg = "375623" if pnl >= 0 else "9C0006"
        
        for j, (col, fmt) in enumerate(zip(PM_COLS, PM_FMTS), 1):
            if col is None:  # Rank column
                val = i + 1
            elif col == 'qty':
                val = int(row[col])
            else:
                val = row[col]
            bold = col in ('net_pnl', 'wi')
            dat(ws, r, j, val, fill=fill, fmt=fmt,
                align=AC if j <= 3 else AR,
                bold=bold, fg=fg if col == 'net_pnl' else "000000")
        ws.row_dimensions[r].height = 16
    
    # Totals row
    tot_row = tbl_hdr_row + 1 + len(pm_df)
    tot_inv = pm_df['invest'].sum()
    tot_exit = pm_df['exit_val'].sum()
    tot_gp = pm_df['gross_pnl'].sum()
    tot_cost = pm_df['cost'].sum()
    tot_np = pm_df['net_pnl'].sum()
    tot_ret = tot_np / tot_inv if tot_inv > EPSILON else 0
    tot_wi = pm_df['wi'].sum()
    
    hdr(ws, tot_row, 1, "TOTAL", bg=F_DARK, merge_to=10)
    tot_vals = [
        (11, tot_wi, FMT_PCT),
        (15, tot_inv, FMT_CASH),
        (16, tot_exit, FMT_CASH),
        (17, tot_gp, FMT_CASH),
        (18, tot_cost, FMT_CASH),
        (19, tot_np, FMT_CASH),
        (20, tot_ret, FMT_PCT),
    ]
    for col, val, fmt in tot_vals:
        c = ws.cell(row=tot_row, column=col, value=val)
        c.font = mkfont(bold=True, color="FFFF00", size=9)
        c.fill = F_DARK
        c.alignment = AR
        c.border = mkborder()
        c.number_format = fmt
    ws.row_dimensions[tot_row].height = 18
    
    # Remaining cells in totals row
    for col in [12, 13, 14]:
        c = ws.cell(row=tot_row, column=col, value="-")
        c.font = mkfont(color="AAAAAA")
        c.fill = F_DARK
        c.alignment = AC
        c.border = mkborder()
    
    # Bottom note
    note_row = tot_row + 2
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=NCOLS)
    c = ws.cell(row=note_row, column=1,
                value=(f"[Note]  DATA BOUNDARY: All signals (beta, ERB, sigma^2, RF) computed using "
                       f"monthly data strictly <= {pm_str}.  "
                       f"Buy/Sell prices sourced from {tm_str} (execution only -- no signal use).  "
                       f"Zero look-ahead bias confirmed. Carry-forward PNL tracking active."))
    c.font = Font(italic=True, color="595959", size=8, name="Arial")
    c.fill = F_GRAY
    c.alignment = AL
    c.border = mkborder()
    ws.row_dimensions[note_row].height = 20
    
    # Correlation Matrix
    corr_start_row = note_row + 2
    ws.merge_cells(start_row=corr_start_row, start_column=1,
                   end_row=corr_start_row, end_column=max(NCOLS, 5))
    c = ws.cell(row=corr_start_row, column=1,
                value="[v]  STOCK CORRELATION MATRIX (Selected Portfolio Only)")
    c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    c.fill = F_TEAL
    c.alignment = AL
    c.border = mkborder()
    
    corr_df = full_correlation_matrices.get(pm_str, pd.DataFrame())
    if not corr_df.empty:
        # Header row for matrix
        for j, sym in enumerate(corr_df.columns, 2):
            hdr(ws, corr_start_row+1, j, sym, bg=F_MID, size=8)
            ws.column_dimensions[gcl(j)].width = 10
        
        # Rows
        for i, (sym, row_data) in enumerate(corr_df.iterrows()):
            curr_row = corr_start_row + 2 + i
            hdr(ws, curr_row, 1, sym, bg=F_MID, size=8, align=AL)
            for j, val in enumerate(row_data, 2):
                # Color-coding
                fill = F_WHT
                if val > 0.8:   fill = F_DRED
                elif val > 0.5: fill = F_ORG
                elif val < 0.2: fill = F_DGRN
                
                dat(ws, curr_row, j, val, fill=fill, fmt=FMT_C2, align=AC)
    
    ws.freeze_panes = f'A{tbl_hdr_row + 1}'
    print(f"   [v]  {sheet_nm}  ({len(pm_df)} stocks)")

# ══════════════════════════════════════════════════════════════════════════════
# BETA METHODOLOGY & MATHEMATICAL EXPLANATION
# ══════════════════════════════════════════════════════════════════════════════

ws_beta = wb.create_sheet("Beta Methodology")

ws_beta.merge_cells('A1:H1')
ws_beta['A1'].value = "MATHEMATICAL METHODOLOGY: PORTFOLIO BETA CALCULATIONS"
ws_beta['A1'].font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
ws_beta['A1'].fill = F_DARK
ws_beta['A1'].alignment = AC
ws_beta.row_dimensions[1].height = 34

# 1. Realized Portfolio Beta
hdr(ws_beta, 3, 1, "1. REALIZED PORTFOLIO BETA (Ex-post)", merge_to=8, bg=F_TEAL)
ws_beta.merge_cells('A4:H8')
beta_realized_text = (
    "The Realized Portfolio Beta measures the actual sensitivity of the portfolio's returns relative to the benchmark "
    "returns over the entire backtest period. It is calculated using the standard OLS (Ordinary Least Squares) "
    "regression slope formula:\n\n"
    "    β_realized = Cov(R_p, R_m) / Var(R_m)\n\n"
    "Where:\n"
    "    • R_p = Actual monthly net returns of the portfolio\n"
    "    • R_m = Monthly returns of the benchmark (NIFTY 50)\n"
    "    • Cov(R_p, R_m) = Covariance between portfolio and market returns\n"
    "    • Var(R_m) = Variance of the market returns"
)
ws_beta['A4'].value = beta_realized_text
ws_beta['A4'].alignment = AL
ws_beta['A4'].font = mkfont(size=10)
ws_beta.row_dimensions[4].height = 120

# 2. Monthly Portfolio Beta
hdr(ws_beta, 10, 1, "2. MONTHLY PORTFOLIO BETA (Ex-ante / Weighted Average)", merge_to=8, bg=F_TEAL)
ws_beta.merge_cells('A11:H15')
beta_exante_text = (
    "The Monthly Portfolio Beta (shown in the dashboard and per-month sheets) is the expected beta at the time of "
    "portfolio construction. It is calculated as the weighted average of the individual stock betas:\n\n"
    "    β_portfolio = Σ (w_i * β_i)\n\n"
    "Where:\n"
    "    • w_i = Weight assigned to stock 'i' in the portfolio\n"
    "    • β_i = Estimated beta of stock 'i' relative to the benchmark (calculated over the lookback window)\n"
    "    • Σ = Summation over all selected stocks in the portfolio"
)
ws_beta['A11'].value = beta_exante_text
ws_beta['A11'].alignment = AL
ws_beta['A11'].font = mkfont(size=10)
ws_beta.row_dimensions[11].height = 100

# 3. Individual Stock Beta
hdr(ws_beta, 17, 1, "3. INDIVIDUAL STOCK BETA ESTIMATION", merge_to=8, bg=F_TEAL)
ws_beta.merge_cells('A18:H22')
beta_stock_text = (
    "Individual stock betas are estimated using historical monthly returns over an adaptive window "
    f"(max {BETA_WINDOW} months, min {MIN_OBS} months). To ensure accuracy, the script identifies the "
    "intersection of available data points between the stock and the benchmark:\n\n"
    "    β_i = Cov(r_i, r_m) / σ_m²\n\n"
    "This calculation uses the intersection-based approach to handle missing data or new listings, "
    "ensuring that the covariance is computed only on aligned historical dates."
)
ws_beta['A18'].value = beta_stock_text
ws_beta['A18'].alignment = AL
ws_beta['A18'].font = mkfont(size=10)
ws_beta.row_dimensions[18].height = 80

# 4. Ex-ante vs. Realized Beta (Why they differ)
hdr(ws_beta, 24, 1, "4. EX-ANTE VS. REALIZED BETA (Why they differ)", merge_to=8, bg=F_TEAL)
ws_beta.merge_cells('A25:H30')
beta_diff_text = (
    "It is common to see a 'Realized Port Beta' (e.g. 0.90) that is higher than the 'Average Ex-ante Beta' (e.g. 0.45). "
    "This happens for several reasons:\n"
    "1. LOOKBACK WINDOW: Ex-ante beta is based on the LAST 60 months of history. Realized beta is based on the "
    "ACTUAL performance DURING the backtest. If correlations rise during the backtest, realized beta will increase.\n"
    "2. MARKET REGIMES: In a bull market or volatile period, even 'low-beta' stocks may move more aggressively "
    "with the index than they did in the past.\n"
    "3. ESTIMATION ERROR: Historical beta is an estimate; actual future sensitivity is never guaranteed.\n"
    "4. DIVERSIFICATION: While individual stocks have low betas, if they are all positively correlated to the "
    "index, the portfolio as a whole will still track the index closely."
)
ws_beta['A25'].value = beta_diff_text
ws_beta['A25'].alignment = AL
ws_beta['A25'].font = mkfont(size=10)
ws_beta.row_dimensions[25].height = 100

for col in range(1, 9):
    ws_beta.column_dimensions[gcl(col)].width = 15

print(f"  [v] Created sheet: Beta Methodology")


# ══════════════════════════════════════════════════════════════════════════════
# CURRENT SELECTION DETAIL (LATEST PORTFOLIO)
# ══════════════════════════════════════════════════════════════════════════════

def add_current_selection_sheet(workbook):
    if not monthly_portfolios:
        return
        
    last_pm_df = monthly_portfolios[-1]
    last_month_str = last_pm_df['port_month'].iloc[0]
    
    ws_last = workbook.create_sheet("Current Selection Detail")
    
    # Title
    ws_last.merge_cells('A1:H1')
    ws_last['A1'].value = f"CURRENT PORTFOLIO SELECTION: {last_month_str}"
    ws_last['A1'].font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
    ws_last['A1'].fill = F_DARK
    ws_last['A1'].alignment = AC
    ws_last.row_dimensions[1].height = 34
    
    # Summary Info
    total_beta = (last_pm_df['wi'] * last_pm_df['beta']).sum()
    n_stocks = len(last_pm_df)
    
    stats = [
        ("Portfolio Month", last_month_str),
        ("Total Stocks", n_stocks),
        ("Portfolio Beta (Ex-ante)", f"{total_beta:.4f}"),
        ("Total Weight", f"{last_pm_df['wi'].sum():.0%}"),
        ("Investment", f"Rs.{INVESTMENT:,.0f}"),
    ]
    
    for i, (lbl, val) in enumerate(stats):
        r = 3 + i
        hdr(ws_last, r, 1, lbl, bg=F_MID, size=9, merge_to=2)
        dat(ws_last, r, 3, val, fill=F_LBLUE, bold=True, align=AC)
    
    # Components Table
    hdr_row = 10
    comp_headers = ["Rank", "Symbol", "Weight (wi)", "Indiv. Beta", "Beta Contribution", "Exp. Return (ERB)", "Resid. Var"]
    comp_cols = [None, 'symbol', 'wi', 'beta', 'beta_cont', 'erb', 'sigma2']
    comp_fmts = [FMT_INT, None, FMT_PCT, FMT_N4, FMT_N4, FMT_PCT, FMT_N6]
    comp_wids = [6, 15, 12, 12, 15, 15, 15]
    
    last_pm_df['beta_cont'] = last_pm_df['wi'] * last_pm_df['beta']
    last_sorted = last_pm_df.sort_values('wi', ascending=False).reset_index(drop=True)
    
    for j, (h, w) in enumerate(zip(comp_headers, comp_wids), 1):
        hdr(ws_last, hdr_row, j, h, bg=F_DARK)
        ws_last.column_dimensions[gcl(j)].width = w
    
    for i, row in last_sorted.iterrows():
        r = hdr_row + 1 + i
        for j, (col, fmt) in enumerate(zip(comp_cols, comp_fmts), 1):
            val = i + 1 if col is None else row[col]
            dat(ws_last, r, j, val, fmt=fmt, align=AL if j == 2 else AR)
            
    print(f"  [v] Created sheet: Current Selection Detail")

add_current_selection_sheet(wb)
add_current_selection_sheet(wb_summary)


# ══════════════════════════════════════════════════════════════════════════════
# SAVE FILES
# ══════════════════════════════════════════════════════════════════════════════

print(f"\n{'='*80}")
print("  SAVING EXCEL FILES")
print(f"{'='*80}\n")


# ==============================================================================
# [Drawdown]  DETAILED DAILY DRAWDOWN ANALYSIS SHEET
# ==============================================================================

def calculate_daily_drawdown(daily_log, monthly_summary):
    df = pd.DataFrame(daily_log)
    
    # 1. Map previous month's compounded wealth to calculate continuous equity curve
    # monthly_summary has 'trade_month' and 'cumulative_ret'
    ms = monthly_summary[['trade_month', 'cumulative_ret']].copy()
    ms['trade_month'] = ms['trade_month'].astype(str)
    ms['prev_cum_ret'] = ms['cumulative_ret'].shift(1).fillna(0)
    
    df['trade_month'] = df['trade_month'].astype(str)
    df = df.merge(ms[['trade_month', 'prev_cum_ret']], on='trade_month', how='left')
    
    # If a month is missing from monthly_summary but present in daily_log, we carry forward the previous prev_cum_ret
    # instead of resetting to 0 (which causes the 88% crash)
    df['prev_cum_ret'] = df['prev_cum_ret'].ffill().fillna(0)
    
    # 2. Continuous Daily Wealth
    # For day 'd' in month 'm', wealth = (1 + prev_month_cum_ret) * (daily_wealth_in_month_m / INVESTMENT)
    df['wealth'] = (1 + df['prev_cum_ret']) * (df['wealth_this_month'] / INVESTMENT)
    
    # 3. Continuous Benchmark Wealth
    df['bench_wealth'] = df['bench_close'] / df['bench_close'].iloc[0]
    
    # 4. Drawdown Calculations
    df['peak'] = df['wealth'].cummax()
    df['drawdown'] = (df['wealth'] - df['peak']) / df['peak']
    
    df['bench_peak'] = df['bench_wealth'].cummax()
    df['bench_drawdown'] = (df['bench_wealth'] - df['bench_peak']) / df['bench_peak']
    
    # 5. Extract Drawdown Periods
    is_dd = df['drawdown'] < -0.0001
    df['dd_group'] = (is_dd != is_dd.shift(1)).cumsum()
    
    dd_periods = []
    
    for grp, group_data in df[is_dd].groupby('dd_group'):
        start_idx = group_data.index[0]
        if start_idx > 0:
            peak_date = df.loc[start_idx - 1, 'date']
            peak_wealth = df.loc[start_idx - 1, 'wealth']
            peak_b_wealth = df.loc[start_idx - 1, 'bench_wealth']
        else:
            peak_date = df.loc[0, 'date']
            peak_wealth = 1.0
            peak_b_wealth = 1.0
            
        trough_idx = group_data['drawdown'].idxmin()
        trough_date = df.loc[trough_idx, 'date']
        max_dd = group_data.loc[trough_idx, 'drawdown']
        
        # Check if recovered
        recovery_date = "Ongoing"
        duration_days = len(group_data)
        
        post_trough = df.loc[trough_idx+1:]
        recovery_slice = post_trough[post_trough['wealth'] >= peak_wealth]
        if not recovery_slice.empty:
            recovery_idx = recovery_slice.index[0]
            recovery_date = df.loc[recovery_idx, 'date']
            duration_days = (recovery_idx - start_idx + 1)
            
        # [FIX] BENCHMARK DRAWDOWN DURING THIS EXACT PERIOD
        # We find the deepest drop the benchmark had BETWEEN peak_date and recovery_date
        period_start_idx = max(0, start_idx - 1)
        period_end_idx = recovery_idx if recovery_date != "Ongoing" else df.index[-1]
        bench_period = df.loc[period_start_idx : period_end_idx]
        
        if not bench_period.empty:
            # We measure the benchmark's own max drawdown during this specific portfolio window
            # This handles cases where the benchmark reached a new peak after the portfolio peak
            # but then dropped significantly before the portfolio recovered.
            b_wealth = bench_period['bench_wealth']
            b_peaks = b_wealth.cummax()
            b_dds = (b_wealth - b_peaks) / b_peaks
            bench_dd = b_dds.min()
        else:
            bench_dd = 0
            
        dd_periods.append({
            'peak_date': peak_date,
            'trough_date': trough_date,
            'recovery_date': recovery_date,
            'duration_days': duration_days,
            'max_dd': max_dd,
            'bench_dd': bench_dd
        })
        
    dd_df = pd.DataFrame(dd_periods)
    if not dd_df.empty:
        dd_df = dd_df.sort_values('max_dd').reset_index(drop=True)
        
    return dd_df, df


def build_drawdown_sheet_v2(workbook, dd_df, df_log, is_daily=False):
    title = "DAILY DRAWDOWN ANALYSIS" if is_daily else "MONTHLY DRAWDOWN ANALYSIS"
    sheet_name = "Daily_Drawdown" if is_daily else "Monthly_Drawdown"
    
    ws = workbook.create_sheet(sheet_name)
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'].value = f"PORTFOLIO {title}"
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    ws['A1'].fill = F_DARK
    ws['A1'].alignment = AC
    
    # Section 1: Top Drawdowns
    hdr(ws, 3, 1, "TOP 5 WORST DRAWDOWN PERIODS", merge_to=6, bg=F_MID)
    dur_label = "Duration (Days)" if is_daily else "Duration (Months)"
    date_label = "Date" if is_daily else "Month"
    for j, h in enumerate(["Rank", f"Peak {date_label}", f"Trough {date_label}", f"Recovery {date_label}", dur_label, "Port Max Drawdown", "Bench Drop in Period"], 1):
        hdr(ws, 4, j, h, bg=F_GRAY, fg="000000")
        
    row_idx = 5
    if not dd_df.empty:
        from openpyxl.formatting.rule import ColorScaleRule
        for i, row in dd_df.head(5).iterrows():
            dat(ws, row_idx, 1, i+1, align=AC)
            p_val = row['peak_date'].strftime('%Y-%m-%d') if is_daily and isinstance(row['peak_date'], pd.Timestamp) else str(row['peak_date'])
            t_val = row['trough_date'].strftime('%Y-%m-%d') if is_daily and isinstance(row['trough_date'], pd.Timestamp) else str(row['trough_date'])
            r_val = row['recovery_date'].strftime('%Y-%m-%d') if is_daily and isinstance(row['recovery_date'], pd.Timestamp) else str(row['recovery_date'])
            
            dat(ws, row_idx, 2, p_val, align=AC)
            dat(ws, row_idx, 3, t_val, align=AC)
            dat(ws, row_idx, 4, r_val, align=AC)
            dat(ws, row_idx, 5, row['duration_days'], fmt=FMT_INT, align=AC)
            dat(ws, row_idx, 6, row['max_dd'], fmt=FMT_PCT, bold=True)
            dat(ws, row_idx, 7, row['bench_dd'], fmt=FMT_PCT)
            row_idx += 1
            
        # Add Color Scale to Drawdown Columns (Reds)
        rule = ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FCB7B9', end_type='max', end_color='FFFFFF')
        ws.conditional_formatting.add(f"F5:G{row_idx-1}", rule)
    else:
        dat(ws, row_idx, 1, "No Drawdowns Found", merge_to=7, align=AC)
        row_idx += 1
        
    row_idx += 2
    
    # Section 2: Summary Stats
    hdr(ws, row_idx, 1, "DRAWDOWN SUMMARY METRICS", merge_to=4, bg=F_MID)
    row_idx += 1
    
    avg_duration = dd_df['duration_days'].mean() if not dd_df.empty else 0
    max_duration = dd_df['duration_days'].max() if not dd_df.empty else 0
    total_dd_time = dd_df['duration_days'].sum() if not dd_df.empty else 0
    total_time_tracked = len(df_log)
    time_underwater = total_dd_time / total_time_tracked if total_time_tracked > 0 else 0
    
    dat(ws, row_idx, 1, f"Average Drawdown Duration ({'Days' if is_daily else 'Months'})", merge_to=3, align=AL, fill=F_LIGHT, bold=True)
    dat(ws, row_idx, 4, avg_duration, fmt=FMT_C2)
    row_idx += 1
    dat(ws, row_idx, 1, f"Maximum Time Underwater ({'Days' if is_daily else 'Months'})", merge_to=3, align=AL, fill=F_LIGHT, bold=True)
    dat(ws, row_idx, 4, max_duration, fmt=FMT_INT)
    row_idx += 1
    dat(ws, row_idx, 1, "% Time Spent Underwater", merge_to=3, align=AL, fill=F_LIGHT, bold=True)
    dat(ws, row_idx, 4, time_underwater, fmt=FMT_PCT)
    
    row_idx += 3
    
    # Section 3: Time Series Log
    hdr(ws, row_idx, 1, f"{'DAILY' if is_daily else 'MONTHLY'} DRAWDOWN LOG", merge_to=3, bg=F_MID)
    row_idx += 1
    hdr(ws, row_idx, 1, date_label, bg=F_GRAY, fg="000000")
    hdr(ws, row_idx, 2, "Portfolio Drawdown", bg=F_GRAY, fg="000000")
    hdr(ws, row_idx, 3, "Benchmark Drawdown", bg=F_GRAY, fg="000000")
    start_log_row = row_idx + 1
    
    for i, row in df_log.iterrows():
        d_val = row['date'].strftime('%Y-%m-%d') if is_daily and isinstance(row['date'], pd.Timestamp) else str(row['trade_month']) if not is_daily else str(row['date'])
        dat(ws, row_idx, 1, d_val, align=AC)
        dat(ws, row_idx, 2, row['drawdown'], fmt=FMT_PCT)
        dat(ws, row_idx, 3, row['bench_drawdown'], fmt=FMT_PCT)
        row_idx += 1
        
    # Color scale for the entire log column
    rule_log = ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FCB7B9', end_type='max', end_color='FFFFFF')
    ws.conditional_formatting.add(f"B{start_log_row}:C{row_idx-1}", rule_log)
        
    for col in range(1, 8):
        ws.column_dimensions[gcl(col)].width = 18

print("Calculating Daily Drawdowns...")
dd_df_daily, df_log_daily = calculate_daily_drawdown(daily_log, monthly_summary)
build_drawdown_sheet_v2(wb, dd_df_daily, df_log_daily, is_daily=True)
build_drawdown_sheet_v2(wb_summary, dd_df_daily, df_log_daily, is_daily=True)

# We still use the monthly calculation for the monthly sheet (which we defined earlier, or we can just reuse the same logic on monthly summary)
# Actually, the user asked to color code the month sheet as well. 
# We can just apply the exact same function to the monthly data!
monthly_log_proxy = []
for i, r in monthly_summary.iterrows():
    monthly_log_proxy.append({
        'date': r['trade_month'],
        'trade_month': r['trade_month'],
        'wealth_this_month': INVESTMENT * (1 + r['return_pct']),
        'bench_close': 1 + r['bench_cum_ret']
    })
dd_df_month, df_log_month = calculate_daily_drawdown(monthly_log_proxy, monthly_summary)
build_drawdown_sheet_v2(wb, dd_df_month, df_log_month, is_daily=False)
build_drawdown_sheet_v2(wb_summary, dd_df_month, df_log_month, is_daily=False)

wb.active = wb["Summary 5Y"]
wb.save(OUTPUT_FILE)
print(f"[save] Main Report Saved -- {OUTPUT_FILE}")

wb_summary.active = wb_summary["Summary 5Y"]
wb_summary.save(SUMMARY_ONLY_FILE)
print(f"[save] Summary-only Report Saved -- {SUMMARY_ONLY_FILE}")

# ══════════════════════════════════════════════════════════════════════════════
# CURRENT PORTFOLIO SELECTION EXPORT (LATEST MONTH)
# ══════════════════════════════════════════════════════════════════════════════

print(f"\n[save] Generating Current Selection: {CURRENT_PORT_FILE}")
last_port = monthly_portfolios[-1]
last_month_name = last_port['trade_month'].iloc[0]

wb_cp = Workbook()
ws_cp = wb_cp.active
ws_cp.title = f"Selection_{last_month_name}"

cp_headers = ["Symbol", "Action", "Qty to Change", "Target Qty", "Prev Qty", "Target Weight %", "Current Price (Rs.)"]
cp_wids = [16, 12, 16, 12, 10, 16, 16]

ws_cp.merge_cells('A1:G1')
ws_cp['A1'].value = f"CURRENT PORTFOLIO SELECTION - {last_month_name}"
ws_cp['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_cp['A1'].fill = F_TEAL
ws_cp['A1'].alignment = AC

for j, (h, w) in enumerate(zip(cp_headers, cp_wids), 1):
    hdr(ws_cp, 3, j, h, bg=F_MID)
    ws_cp.column_dimensions[gcl(j)].width = w

for i, row in last_port.iterrows():
    r = i + 4
    action = row['action_needed']
    fill = F_DGRN if "BUY" in action else (F_DRED if "SELL" in action else F_LIGHT)
    
    dat(ws_cp, r, 1, row['symbol'], align=AL, bold=True)
    dat(ws_cp, r, 2, action.split()[0], fill=fill, align=AC, bold=True)
    dat(ws_cp, r, 3, int(abs(row['delta_qty'])), fill=fill, align=AC, fmt=FMT_INT)
    dat(ws_cp, r, 4, int(row['qty']), align=AC, fmt=FMT_INT)
    dat(ws_cp, r, 5, int(row['prev_qty']), align=AC, fmt=FMT_INT)
    dat(ws_cp, r, 6, row['wi'], fmt=FMT_PCT)
    dat(ws_cp, r, 7, row['buy_price'], fmt=FMT_C2)

wb_cp.save(CURRENT_PORT_FILE)
print(f"[save] Current Portfolio Instructions Saved -- {CURRENT_PORT_FILE}")

print(f"\n{'='*80}")
print("[v] UNIFIED BACKTESTER COMPLETE!")
print(f"{'='*80}")
