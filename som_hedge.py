import os
import warnings
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime, time, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

warnings.filterwarnings("ignore")

# ==============================================================================
# [Config] CONFIGURATION
# ==============================================================================

# Input folders and files
STOCKS_FOLDER    = os.environ.get("STOCKS_FOLDER", "nifty500_host")
BOND_FILE        = "India 1-Year Bond Yield Historical Data.csv"
BENCHMARK_FILE   = os.environ.get("BENCHMARK_FILE", "NIFTY500_1d.csv")
NIFTY_SPOT_DIR   = "spot_parquet"
NIFTY_FUTURES_DIR = "futures_parquet"

# Output files
OUTPUT_FILE      = os.environ.get("OUTPUT_FILE", "Hedge_nifty50.xlsx")
DEEP_DIVE_FILE   = os.environ.get("DEEP_DIVE_FILE", "Hedge_Institutional_Deep_Dive_nifty50.xlsx")

# Backtest period
START_MONTH      = "2021-04"
# Analysis ends at the previous month to allow current month to be the 'Live Performance' month
END_MONTH        = datetime.now().strftime("%Y-%m") 

# Parameters
MAX_WEIGHT       = 0.10     # Max 10% per stock
INITIAL_CAPITAL  = 10000000 # 1 Cr
COST_PER_TRADE   = 0.002    # 0.2% churning cost
LOT_SIZE         = 65       # Nifty lot size

# ==============================================================================
# [Core] ADVANCED ANALYTICS UTILITIES
# ==============================================================================

def calculate_supertrend(df, period=10, multiplier=3):
    """Calculate Supertrend indicator (1 = Bullish, -1 = Bearish)"""
    if len(df) < period: return pd.Series(0, index=df.index)
    high, low, close = df.get('high', df['close']), df.get('low', df['close']), df['close']
    tr = pd.concat([high - low, abs(high - close.shift()), abs(low - close.shift())], axis=1).max(axis=1)
    atr = tr.rolling(period).mean()
    hl2 = (high + low) / 2
    upper_band, lower_band = hl2 + (multiplier * atr), hl2 - (multiplier * atr)
    
    final_upper, final_lower = upper_band.copy(), lower_band.copy()
    for i in range(1, len(df)):
        final_upper.iloc[i] = min(upper_band.iloc[i], final_upper.iloc[i-1]) if close.iloc[i-1] <= final_upper.iloc[i-1] else upper_band.iloc[i]
        final_lower.iloc[i] = max(lower_band.iloc[i], final_lower.iloc[i-1]) if close.iloc[i-1] >= final_lower.iloc[i-1] else lower_band.iloc[i]
            
    st = pd.Series(1, index=df.index)
    trend = 1
    for i in range(1, len(df)):
        if trend == 1:
            if close.iloc[i] < final_lower.iloc[i]: trend, st.iloc[i] = -1, -1
        else:
            if close.iloc[i] > final_upper.iloc[i]: trend, st.iloc[i] = 1, 1
            else: st.iloc[i] = -1
    return st

def calculate_ema(df, period=40):
    return df['close'].ewm(span=period, adjust=False).mean()

def calculate_xirr(returns, initial_capital):
    """Simple XIRR approximation for monthly returns"""
    if returns.empty: return 0.0
    total_ret = (1 + returns).prod() - 1
    years = len(returns) / 12.0
    return ((1 + total_ret) ** (1/years)) - 1 if total_ret > -1 and years > 0 else 0.0

def resample_to_monthly(daily_df):
    """Matches SOM.py resample logic exactly"""
    df = daily_df.copy()
    if 'Date' in df.columns: df.rename(columns={'Date': 'time'}, inplace=True)
    df.columns = [c.lower() for c in df.columns]
    df['time'] = pd.to_datetime(df['time'], dayfirst=True)
    df = df.set_index('time').sort_index()
    
    agg_dict = {'open': 'first', 'high': 'max', 'low': 'min', 'close': 'last'}
    agg_dict = {k: v for k, v in agg_dict.items() if k in df.columns}
    
    monthly = df.resample('ME').agg(agg_dict)
    monthly.index = monthly.index.to_period('M')
    monthly['return'] = monthly['close'].pct_change().fillna(0)
    return monthly

def get_rolling_returns(returns, window_months):
    if len(returns) < window_months: return np.nan
    rolling_rets = (1 + returns).rolling(window=window_months).apply(np.prod, raw=True) - 1
    return rolling_rets.mean()

def get_advanced_metrics(returns, benchmark_returns=None, rf_annual=0.06):
    """Institutional-grade performance metrics"""
    if returns.empty: return {}
    
    # 1. Basic Returns
    total_ret = (1 + returns).prod() - 1
    years = len(returns) / 12.0
    cagr = ((1 + total_ret) ** (1/years)) - 1 if total_ret > -1 and years > 0 else -1.0
    
    # 2. Risk Metrics
    vol = returns.std() * np.sqrt(12)
    downside_returns = returns[returns < 0]
    downside_vol = np.sqrt(np.mean(downside_returns**2)) * np.sqrt(12) if len(downside_returns) > 0 else 0.001
    
    cum_ret = (1 + returns).cumprod()
    peak = cum_ret.cummax()
    drawdown = (cum_ret - peak) / peak
    mdd = drawdown.min()
    
    # Drawdown Duration
    is_in_dd = drawdown < 0
    if not is_in_dd.any():
        mdd_duration = 0
    else:
        dd_streak = is_in_dd.astype(int).groupby(is_in_dd.eq(0).cumsum()).cumsum()
        mdd_duration = dd_streak.max()

    # VaR / CVaR (95% & 99%)
    var_95 = np.percentile(returns, 5)
    var_99 = np.percentile(returns, 1)
    cvar_95 = returns[returns <= var_95].mean() if len(returns[returns <= var_95]) > 0 else var_95
    cvar_99 = returns[returns <= var_99].mean() if len(returns[returns <= var_99]) > 0 else var_99

    # 3. Ratios
    sharpe = (cagr - rf_annual) / vol if vol > 0 else 0
    sortino = (cagr - rf_annual) / downside_vol if downside_vol > 0 else 0
    calmar = cagr / abs(mdd) if abs(mdd) > 0 else 0
    
    # Alpha / Information Ratio
    alpha, info_ratio = 0.0, 0.0
    if benchmark_returns is not None:
        common_idx = returns.index.intersection(benchmark_returns.index)
        if len(common_idx) > 0:
            ret_sub, bench_sub = returns.loc[common_idx], benchmark_returns.loc[common_idx]
            bench_total_ret = (1 + bench_sub).prod() - 1
            bench_cagr = ((1 + bench_total_ret) ** (1/years)) - 1
            alpha = cagr - bench_cagr
            tracking_error = (ret_sub - bench_sub).std() * np.sqrt(12)
            info_ratio = alpha / tracking_error if tracking_error > 0 else 0

    # 4. Trade & Execution Metrics
    win_rate = len(returns[returns > 0]) / len(returns)
    avg_gain = returns[returns > 0].mean() if not returns[returns > 0].empty else 0
    avg_loss = returns[returns < 0].mean() if not returns[returns < 0].empty else 0
    profit_factor = abs(returns[returns > 0].sum() / returns[returns < 0].sum()) if returns[returns < 0].sum() != 0 else 10.0
    expectancy = (win_rate * avg_gain) + ((1 - win_rate) * avg_loss)
    
    return {
        "CAGR": cagr, "XIRR": calculate_xirr(returns, INITIAL_CAPITAL), "Abs Return": total_ret, "Alpha vs Bench": alpha,
        "Volatility": vol, "Downside Dev": downside_vol, "Sharpe": sharpe,
        "Sortino": sortino, "Calmar": calmar, "Max Drawdown": mdd, "DD Duration (M)": mdd_duration,
        "VaR 95%": var_95, "VaR 99%": var_99, "CVaR 95%": cvar_95, "CVaR 99%": cvar_99, "Info Ratio": info_ratio,
        "Win Rate": win_rate, "Profit Factor": profit_factor, "Expectancy": expectancy,
        "Avg Gain": avg_gain, "Avg Loss": avg_loss,
        "Rolling 1Y": get_rolling_returns(returns, 12),
        "Rolling 3Y": get_rolling_returns(returns, 36),
        "Best Month": returns.max(), "Worst Month": returns.min()
    }

def calculate_ex_ante_metrics(selected_stocks, port_month, all_stocks, rf_annual, beta_window=60):
    """Calculate ex-ante risk metrics using daily returns for robustness"""
    if not selected_stocks: return 0.0, 0.0, 0.0, 0.0
    weights = np.array([s['weight'] for s in selected_stocks])
    returns = np.array([s['annual_ret'] for s in selected_stocks])
    port_exp_ret = np.sum(weights * returns)
    port_beta = np.sum(weights * np.array([s['beta'] for s in selected_stocks]))
    
    # Use Daily Returns for covariance matrix to handle new stocks and non-overlapping history
    returns_list = []
    # For PeriodIndex, port_month is already the month we need
    # Comparison should be done against the PeriodIndex directly
    for s in selected_stocks:
        sym = s['ticker']
        if sym in all_stocks_daily:
            s_daily = all_stocks_daily[sym]
            end_date = port_month.to_timestamp(how='end')
            s_hist = s_daily[s_daily.index <= end_date]
            if len(s_hist) > 5:
                # Get last beta_window * 21 trading days (approx)
                s_rets = s_hist['close'].pct_change().dropna().tail(beta_window * 21)
                if len(s_rets) > 0:
                    returns_list.append(s_rets.rename(sym))
    
    port_vol = 0.0
    if len(returns_list) > 0:
        ret_matrix = pd.concat(returns_list, axis=1).dropna()
        if not ret_matrix.empty and len(ret_matrix) >= 5:
            cov_mat_annual = ret_matrix.cov() * 252 # Annualize from daily
            port_variance = weights.T @ cov_mat_annual.values @ weights
            port_vol = np.sqrt(max(0, port_variance))
    
    ex_ante_sr = (port_exp_ret - rf_annual) / port_vol if port_vol > 1e-10 else 0.0
    return port_exp_ret, port_vol, ex_ante_sr, port_beta

# ==============================================================================
# [Data] LOADING UTILITIES
# ==============================================================================

def load_intraday_data(folder_path, start_date=None, end_date=None):
    if not os.path.exists(folder_path): 
        print(f"    [!] Warning: Folder not found: {folder_path}")
        return pd.DataFrame()
    files = sorted([f for f in os.listdir(folder_path) if f.endswith('.parquet')])
    if not files:
        print(f"    [!] Warning: No parquet files found in {folder_path}")
        return pd.DataFrame()
        
    dfs = []
    # print(f"    [*] Loading intraday data from {folder_path}...")
    for f in files:
        dt_str = f.replace('.parquet', '')
        try:
            if start_date or end_date:
                dt = pd.to_datetime(dt_str)
                if start_date and dt < start_date: continue
                if end_date and dt > end_date: continue
            
            tmp = pd.read_parquet(os.path.join(folder_path, f))
            if tmp.empty: continue
            
            # Use 'Date' and 'Time' explicitly as they appear in the file
            if 'Time' in tmp.columns:
                tmp['datetime'] = pd.to_datetime(tmp['Date'].astype(str) + ' ' + tmp['Time'].astype(str))
            else:
                tmp['datetime'] = pd.to_datetime(tmp['Date'].astype(str))
            
            tmp.set_index('datetime', inplace=True)
            tmp.columns = [c.lower() for c in tmp.columns]
            dfs.append(tmp)
        except: pass
    
    if not dfs: return pd.DataFrame()
    df_combined = pd.concat(dfs).sort_index()
    # Ensure DatetimeIndex for robustness
    if not isinstance(df_combined.index, pd.DatetimeIndex):
        df_combined.index = pd.to_datetime(df_combined.index)
    return df_combined

def get_mtf_signals(spot_df):
    if spot_df.empty: return pd.DataFrame()
    resampled = {tf: spot_df.resample(r).agg({'open':'first','high':'max','low':'min','close':'last'}).dropna() for tf, r in [('15m','15min'),('30m','30min'),('1h','60min'),('1d','D')]}
    signals = pd.DataFrame(index=spot_df.index)
    for tf, df_tf in resampled.items():
        if df_tf.empty:
            signals[tf] = 0
            continue
        st = calculate_supertrend(df_tf)
        signals[tf] = st.shift(1).reindex(spot_df.index, method='ffill').fillna(0)
    signals['is_bearish_aligned'] = (signals['15m'] == -1) & (signals['30m'] == -1) & (signals['1h'] == -1) & (signals['1d'] == -1)
    signals['is_bullish_aligned'] = (signals['15m'] == 1) & (signals['30m'] == 1) & (signals['1h'] == 1) & (signals['1d'] == 1)
    return signals

# ==============================================================================
# [Logic] DEFENSE LAYER DEFINITIONS (IMPLEMENTATION DETAILS)
# ==============================================================================
# 1. Base SIM: Treynor-Black Optimization. Calculated Beta/ERB, applied cut-off (Ci), 
#    and capped individual weights at 10% for diversification.
# 2. ST Filter: Momentum Defense. Applied monthly Supertrend (10,3); only kept 
#    stocks with Bullish (1) status to avoid intermediate downtrends.
# 3. EMA Filter: Trend Defense. Applied 40-period EMA; only retained stocks trading 
#    above their primary trend average.
# 4. COMBO Filter: Double-Validation. Stocks must pass BOTH ST and EMA-40 filters, 
#    ensuring high-conviction momentum alignment.
# 5. ULTRA Layer: Concentrated Alpha. Synced with COMBO selection as a base for 
#    aggressive hedging when market breadth deteriorates.
# 6. COMBO + Hedge: Active Beta Hedging. Shorts Nifty Futures proportional to 
#    portfolio beta when 15m/30m/1h/1d Nifty Supertrends all turn Bearish.
# 7. ULTRA Defense: Aggressive Multiplier Hedge. Applies a 2.0x Beta multiplier 
#    to Nifty shorts when market breadth (Combo/Base ratio) falls below 50%.
# ==============================================================================

# ==============================================================================
# [Logic] THE BACKTESTER
# ==============================================================================

print("\n" + "="*80)
print("  SOM_HEDGE: INSTITUTIONAL ANALYTICS ENGINE")
print("="*80)

bench_df = pd.read_csv(BENCHMARK_FILE, parse_dates=['Date'], index_col='Date').sort_index()
bench_df.columns = [c.lower() for c in bench_df.columns]
bench_monthly = bench_df['close'].resample('ME').last().pct_change().dropna()
bench_monthly.index = bench_monthly.index.to_period('M')

bond_df = pd.read_csv(BOND_FILE)
bond_df['Date'] = pd.to_datetime(bond_df['Date'], errors='coerce', dayfirst=True)
bond_df.set_index('Date', inplace=True)
rf_monthly = bond_df['Price'].resample('ME').last() / 100 / 12
rf_monthly.index = rf_monthly.index.to_period('M')
# Ensure we have data for the full range
rf_monthly = rf_monthly.ffill()

nifty_spot = load_intraday_data(NIFTY_SPOT_DIR)
nifty_futures = load_intraday_data(NIFTY_FUTURES_DIR)
mtf_signals = get_mtf_signals(nifty_spot)

all_stocks = {}
all_stocks_daily = {}
stock_files = list(Path(STOCKS_FOLDER).glob("*.csv"))
for i, sf in enumerate(stock_files):
    if i % 100 == 0: print(f"    [*] Loading stocks: {i}/{len(stock_files)}...", end='\r')
    try:
        sdf = pd.read_csv(sf)
        if not sdf.empty:
            mdf = resample_to_monthly(sdf)
            all_stocks[sf.stem] = mdf
            # For daily fallback, ensure columns are lowercased
            sdf.columns = [c.lower() for c in sdf.columns]
            date_col = 'date' if 'date' in sdf.columns else 'time'
            if date_col in sdf.columns:
                sdf[date_col] = pd.to_datetime(sdf[date_col], dayfirst=True)
                all_stocks_daily[sf.stem] = sdf.set_index(date_col).sort_index()
    except: pass

all_port_months = pd.period_range(start=START_MONTH, end=END_MONTH, freq='M')
monthly_summary = []

# Churning tracking
prev_layer_stocks = {layer: set() for layer in ['Base', 'ST', 'EMA', 'COMBO', 'ULTRA']}
churn_history = []
added_stocks_total = {layer: {} for layer in ['Base', 'ST', 'EMA', 'COMBO', 'ULTRA']}
removed_stocks_total = {layer: {} for layer in ['Base', 'ST', 'EMA', 'COMBO', 'ULTRA']}

prev_weights = {layer: {} for layer in ['Base', 'ST', 'EMA', 'COMBO', 'ULTRA']}
# Sizing always stays at INITIAL_CAPITAL (Fixed Investment logic)
equities = {layer: INITIAL_CAPITAL for layer in ['Base', 'ST', 'EMA', 'COMBO', 'ULTRA', 'COMBO_HEDGE', 'ULTRA_HEDGE']}

EPSILON, BETA_WINDOW, MIN_OBS = 1e-10, 60, 2
portfolio_audit_log = []

for port_month in all_port_months:
    # Matches SOM.py: Trade in the NEXT month using data UP TO the current port_month
    trade_month = port_month + 1
    
    print(f"  [{port_month.strftime('%Y-%m')}] -> Trading {trade_month.strftime('%Y-%m')}...")

    
    bench_hist = bench_monthly[bench_monthly.index <= port_month].tail(BETA_WINDOW)
    if len(bench_hist) < MIN_OBS: continue
    bench_var_annual = bench_hist.var() * 12
    avg_rf_month = rf_monthly.reindex([port_month], method='ffill').iloc[0] * 12

    candidates = []
    # Benchmark Daily for beta fallback
    bench_daily_ret = bench_df['close'].pct_change() # Don't dropna here to match SOM.py

    for ticker, mdf in all_stocks.items():
        # Matches SOM.py: Use pre-processed monthly returns
        stock_hist = mdf[mdf.index <= port_month]
        stock_window = stock_hist.tail(BETA_WINDOW)
        common_idx = stock_window.index.intersection(bench_hist.index)
        
        # Matches SOM.py: Only use daily fallback if monthly observations are < MIN_OBS (2)
        use_daily = len(common_idx) < MIN_OBS 
        
        if use_daily:
            # ✅ DAILY FALLBACK
            s_daily = all_stocks_daily.get(ticker, pd.DataFrame())
            if s_daily.empty: continue
            end_date = port_month.to_timestamp(how='end')
            s_daily_hist = s_daily[s_daily.index <= end_date].tail(BETA_WINDOW * 21)
            if len(s_daily_hist) < 5: continue
            
            s_returns = s_daily_hist['close'].pct_change().dropna()
            m_returns = bench_daily_ret.reindex(s_returns.index).dropna()
            
            common_idx_daily = s_returns.index.intersection(m_returns.index)
            if len(common_idx_daily) < 5: continue
            
            s_rets_arr = s_returns.loc[common_idx_daily].values
            m_rets_arr = m_returns.loc[common_idx_daily].values
            ann_factor = 252
        else:
            s_rets_arr = stock_window.loc[common_idx, 'return'].values
            m_rets_arr = bench_hist.loc[common_idx].values
            ann_factor = 12

        cov = np.cov(s_rets_arr, m_rets_arr, ddof=1)
        beta = cov[0, 1] / cov[1, 1]
        if beta <= 0 or pd.isna(beta): continue

        # Survival check (Matches SOM.py)
        if trade_month in mdf.index:
            trade_row = mdf.loc[trade_month]
            buy_px = trade_row.get('open', np.nan)
            sell_px = trade_row.get('close', np.nan)
            if pd.isna(buy_px) or buy_px <= 0: buy_px = stock_hist['close'].iloc[-1]
            if pd.isna(sell_px): sell_px = buy_px
            trade_data_exists = True
        else:
            buy_px = stock_hist['close'].iloc[-1]
            sell_px = buy_px
            trade_data_exists = False
            
        if pd.isna(buy_px) or buy_px <= 0: continue

        # Residual Variance (SIM Method)
        ann_ret = s_rets_arr.mean() * ann_factor
        alpha_val = ann_ret - beta * m_rets_arr.mean() * ann_factor
        sigma2 = np.var(s_rets_arr - (alpha_val/ann_factor + beta * m_rets_arr), ddof=1) * ann_factor
        
        if pd.isna(sigma2) or sigma2 < EPSILON: continue

        candidates.append({
            'ticker': ticker, 'beta': beta, 'annual_ret': ann_ret, 
            'sigma2': sigma2, 'erb': (ann_ret - avg_rf_month)/beta,
            'buy_px': buy_px, 'sell_px': sell_px, 'trade_data_exists': trade_data_exists
        })

    if not candidates: continue
    candidates = sorted(candidates, key=lambda x: x['erb'], reverse=True)
    cum_A, cum_H, ci_list = 0, 0, []
    avg_rf_month = rf_monthly.reindex([port_month], method='ffill').iloc[0] * 12
    for c in candidates:
        cum_A += ((c['annual_ret'] - avg_rf_month) * c['beta']) / c['sigma2']
        cum_H += (c['beta'] ** 2) / c['sigma2']
        ci_list.append((bench_var_annual * cum_A) / (1 + bench_var_annual * cum_H))

    best_ci = max(ci_list)
    selected = candidates[:ci_list.index(best_ci)+1]
    for s in selected: s['z'] = max(0, (s['beta'] / s['sigma2']) * (s['erb'] - best_ci))
    total_z = sum(s['z'] for s in selected)
    if total_z < EPSILON: continue
    
    w = np.array([s['z']/total_z for s in selected], dtype=float)
    for _ in range(1000):
        over = w > MAX_WEIGHT
        if not over.any(): break
        excess = (w[over] - MAX_WEIGHT).sum()
        w[over], under = MAX_WEIGHT, ~over
        if not under.any(): break
        w[under] += excess * (w[under] / w[under].sum())
    weights = w / w.sum()

    for s, w in zip(selected, weights): s['weight'] = w
    selected = [s for s in selected if s['weight'] > 0]

    layers_current = {'Base': {s['ticker'] for s in selected}}
    st_kept, ema_kept, combo_kept = [], [], []
    for s in selected:
        hist = all_stocks[s['ticker']][all_stocks[s['ticker']].index <= port_month].tail(50)
        p_st = calculate_supertrend(hist).iloc[-1] == 1 if not hist.empty else False
        p_ema = hist['close'].iloc[-1] > calculate_ema(hist, 40).iloc[-1] if not hist.empty else False
        if p_st: st_kept.append(s)
        if p_ema: ema_kept.append(s)
        if p_st and p_ema: combo_kept.append(s)
    
    layers_current.update({'ST': {s['ticker'] for s in st_kept}, 'EMA': {s['ticker'] for s in ema_kept}, 'COMBO': {s['ticker'] for s in combo_kept}, 'ULTRA': {s['ticker'] for s in combo_kept}})
    
    churn_row = {'Month': port_month.strftime('%Y-%m')}
    port_exp_ret, port_vol, ex_ante_sr, port_beta = calculate_ex_ante_metrics(selected, port_month, all_stocks, avg_rf_month, BETA_WINDOW)
    churn_row['Ex_Ante_Sharpe'] = ex_ante_sr
    churn_row['Port_Beta'] = port_beta
    churn_row['Stock_Count'] = len(selected)

    for l_name, c_stocks in layers_current.items():
        added = c_stocks - prev_layer_stocks[l_name]
        removed = prev_layer_stocks[l_name] - c_stocks
        churn_row[f'{l_name} Add'], churn_row[f'{l_name} Rem'] = len(added), len(removed)
        for t in added: added_stocks_total[l_name][t] = added_stocks_total[l_name].get(t, 0) + 1
        for t in removed: removed_stocks_total[l_name][t] = removed_stocks_total[l_name].get(t, 0) + 1
    churn_history.append(churn_row)

    tm_start, tm_end = trade_month.to_timestamp(), trade_month.to_timestamp(how='end')
    layer_returns = {l: 0.0 for l in layers_current.keys()}
    curr_weights = {l: {s['ticker']: 0.0 for s in selected} for l in layers_current.keys()}
    
    for s in selected:
        ticker = s['ticker']
        curr_weights['Base'][ticker] = s['weight']
        if ticker in layers_current['ST']: curr_weights['ST'][ticker] = s['weight']
        if ticker in layers_current['EMA']: curr_weights['EMA'][ticker] = s['weight']
        if ticker in layers_current['COMBO']: curr_weights['COMBO'][ticker] = s['weight']
        if ticker in layers_current['ULTRA']: curr_weights['ULTRA'][ticker] = s['weight']

    all_tickers_to_process = set()
    # --- QUANTITY-BASED PNL CALCULATION (Matches SOM.py) ---
    layer_pnl_rupees = {l: 0.0 for l in layers_current.keys()}
    
    all_tickers_to_process = set()
    for l in layers_current.keys():
        all_tickers_to_process.update(curr_weights[l].keys())
        all_tickers_to_process.update(prev_weights[l].keys())
    
    selected_tickers = {s['ticker'] for s in selected}

    for ticker in all_tickers_to_process:
        mdf = all_stocks.get(ticker)
        if mdf is None: continue
        
        # Prices from SOM.py logic
        prev_close = mdf[mdf.index <= port_month]['close'].iloc[-1] if not mdf[mdf.index <= port_month].empty else 0
        if trade_month in mdf.index:
            buy_px = mdf.loc[trade_month, 'open']
            if pd.isna(buy_px) or buy_px <= 0: buy_px = prev_close
            sell_px = mdf.loc[trade_month, 'close']
        else:
            buy_px = prev_close
            sell_px = buy_px

        for l in layers_current.keys():
            w_prev = prev_weights[l].get(ticker, 0.0)
            w_curr = curr_weights[l].get(ticker, 0.0)
            
            # Quantities (Matches SOM.py floor logic)
            q_p = int(np.floor((INITIAL_CAPITAL * w_prev) / prev_close)) if prev_close > 0 else 0
            q_c = int(np.floor((INITIAL_CAPITAL * w_curr) / buy_px)) if buy_px > 0 else 0
            
            # PNL Components (Gap + Current)
            gap_pnl = (buy_px - prev_close) * q_p
            curr_pnl = (sell_px - buy_px) * q_c
            layer_pnl_rupees[l] += (gap_pnl + curr_pnl)

        # Audit Log Entry
        if ticker in selected_tickers:
            s = [s for s in selected if s['ticker'] == ticker][0]
            curr_qty = int(np.floor((INITIAL_CAPITAL * s['weight']) / buy_px))
            prev_qty = 0
            prev_month_str = (port_month - 1).strftime('%Y-%m')
            prev_audit = [a for a in portfolio_audit_log if a['Month'] == prev_month_str and a['Symbol'] == ticker]
            if prev_audit: prev_qty = prev_audit[0].get('Qty', 0)
            
            delta_qty = curr_qty - prev_qty
            action = "HOLD"
            if delta_qty > 0: action = f"BUY {int(delta_qty)}"
            elif delta_qty < 0: action = f"SELL {int(abs(delta_qty))}"

            l_wts = {l: curr_weights[l].get(ticker, 0.0) for l in ['Base', 'ST', 'EMA', 'COMBO', 'ULTRA']}
            portfolio_audit_log.append({
                'Month': port_month.strftime('%Y-%m'), 
                'Analysis_Month': port_month.strftime('%Y-%m'),
                'Trade_Month': trade_month.strftime('%Y-%m'),
                'Symbol': ticker, 'Beta': s['beta'], 
                'ERB': s['erb'], 'Weight': s['weight'], 
                'ST_Wt': l_wts['ST'], 'EMA_Wt': l_wts['EMA'], 'COMBO_Wt': l_wts['COMBO'], 'ULTRA_Wt': l_wts['ULTRA'],
                'Qty': curr_qty, 'Prev_Qty': prev_qty, 'Delta_Qty': delta_qty, 'Action': action,
                'Status': 'Added' if ticker not in prev_layer_stocks['Base'] else 'Remained',
                'Buy_Px': buy_px, 'Sell_Px': sell_px, 'Return': (sell_px/buy_px)-1 if buy_px > 0 else 0.0
            })

    # Transaction Costs (Quantity-Based Turnover)
    for l in layers_current.keys():
        # Turnover value = sum of absolute quantity changes * price
        txn_value = 0
        for t in set(curr_weights[l]) | set(prev_weights[l]):
            w_c, w_p = curr_weights[l].get(t, 0), prev_weights[l].get(t, 0)
            mdf = all_stocks.get(t)
            if mdf is None or mdf.empty: continue
            p_c = mdf[mdf.index <= port_month]['close'].iloc[-1]
            b_p = mdf.loc[trade_month, 'open'] if trade_month in mdf.index else p_c
            if pd.isna(b_p) or b_p <= 0: b_p = p_c
            
            q_curr = int(np.floor((INITIAL_CAPITAL * w_c) / b_p)) if b_p > 0 else 0
            q_prev = int(np.floor((INITIAL_CAPITAL * w_p) / p_c)) if p_c > 0 else 0
            txn_value += abs(q_curr - q_prev) * b_p
            
        txn_cost_rupees = txn_value * COST_PER_TRADE
        layer_returns[l] = (layer_pnl_rupees[l] - txn_cost_rupees) / INITIAL_CAPITAL
        prev_weights[l] = curr_weights[l].copy()

    # --- HEDGING (Fixed Investment Style) ---
    h_pnl, h_pnl_u, is_h, entry_f, port_beta = 0, 0, False, 0, sum(s['weight']*s['beta'] for s in selected)
    
    # Robust slicing using .loc for time-based indexing
    if not mtf_signals.empty:
        m_sig = mtf_signals.loc[tm_start:tm_end]
    else:
        m_sig = pd.DataFrame()
        
    if not nifty_futures.empty:
        m_fut = nifty_futures.loc[tm_start:tm_end]
    else:
        m_fut = pd.DataFrame()
    multiplier = 2.0 if (len(combo_kept)/len(selected) < 0.5) else 1.0
    
    # Sizing for hedging also uses INITIAL_CAPITAL
    lots = 0
    lots_u = 0
    
    for t in m_sig.index:
        if not is_h and m_sig.loc[t, 'is_bearish_aligned'] and t in m_fut.index:
            entry_f, is_h = m_fut.loc[t, 'close'], True
            if isinstance(entry_f, pd.Series): entry_f = entry_f.iloc[0]
            lots = round((INITIAL_CAPITAL * port_beta) / (entry_f * LOT_SIZE)) if entry_f > 0 else 0
            lots_u = round((INITIAL_CAPITAL * port_beta * multiplier) / (entry_f * LOT_SIZE)) if entry_f > 0 else 0
        elif is_h and m_sig.loc[t, 'is_bullish_aligned'] and t in m_fut.index:
            exit_f = m_fut.loc[t, 'close']
            if isinstance(exit_f, pd.Series): exit_f = exit_f.iloc[0]
            h_pnl += (entry_f - exit_f) * LOT_SIZE * lots
            h_pnl_u += (entry_f - exit_f) * LOT_SIZE * lots_u
            is_h = False
    if is_h and not m_fut.empty:
        pts = (entry_f - m_fut['close'].iloc[-1]) * LOT_SIZE
        h_pnl += pts * lots; h_pnl_u += pts * lots_u

    ch_ret = layer_returns['COMBO'] + (h_pnl / INITIAL_CAPITAL)
    uh_ret = layer_returns['ULTRA'] + (h_pnl_u / INITIAL_CAPITAL)
    
    summary_row = {
        'Analysis_Month': port_month.strftime('%Y-%m'), 
        'Month': trade_month.strftime('%Y-%m'), # Use Trade_Month as 'Month' for better alignment
        'Trade_Month': trade_month.strftime('%Y-%m'),
        'Port_Beta': port_beta,
        'Ex_Ante_Sharpe': ex_ante_sr,
        'Stock_Count': len(selected),
        'Added': len(layers_current['Base'] - prev_layer_stocks['Base']),
        'Removed': len(prev_layer_stocks['Base'] - layers_current['Base']),
        'Base': layer_returns['Base'], 'ST': layer_returns['ST'], 'EMA': layer_returns['EMA'], 
        'COMBO': layer_returns['COMBO'], 'ULTRA': layer_returns['ULTRA'], 
        'COMBO_HEDGE': ch_ret, 'ULTRA_HEDGE': uh_ret, 'Bench': bench_monthly.get(trade_month, 0)
    }
    monthly_summary.append(summary_row)
    # Update state for next month's comparison
    for l_name in layers_current.keys():
        prev_layer_stocks[l_name] = layers_current[l_name].copy()

# ==============================================================================
# [Report] EXCEL GENERATION
# ==============================================================================

print("\n[*] Finalizing Professional Report...")
df_sum = pd.DataFrame(monthly_summary).set_index('Month')
df_churn = pd.DataFrame(churn_history)
wb = Workbook()

THIN = Side(border_style="thin", color="000000")
BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
WHITE_FONT = Font(color="FFFFFF", bold=True)

def style_cell(cell, fill=None, font=None, border=BRD, align="center"):
    if fill: cell.fill = fill
    if font: cell.font = font
    if border: cell.border = border
    cell.alignment = Alignment(horizontal=align, vertical="center")

# 1. Dashboard
ws_dash = wb.active
ws_dash.title = "Executive_Dashboard"
headers = ["Metric", "Base SIM", "ST Filter", "EMA Filter", "COMBO Filter", "ULTRA Layer", "COMBO+Hedge", "ULTRA Defense"]
layers = ['Base', 'ST', 'EMA', 'COMBO', 'ULTRA', 'COMBO_HEDGE', 'ULTRA_HEDGE']

for j, h in enumerate(headers, 1):
    c = ws_dash.cell(1, j, h); style_cell(c, HDR_FILL, WHITE_FONT)
    ws_dash.column_dimensions[gcl(j)].width = 20

# Calculate average RF for the backtest period to use in summary metrics
avg_rf_period = (rf_monthly[rf_monthly.index.isin(all_port_months)].mean() * 12) if not rf_monthly.empty else 0.06
metrics_results = {l: get_advanced_metrics(df_sum[l], df_sum['Bench'], rf_annual=avg_rf_period) for l in layers}
avg_ex_ante = df_sum['Ex_Ante_Sharpe'].mean()
for l in layers:
    metrics_results[l]['Avg Ex-Ante Sharpe'] = avg_ex_ante

metric_names = list(metrics_results['Base'].keys())

for i, m_name in enumerate(metric_names, 2):
    c = ws_dash.cell(i, 1, m_name); style_cell(c, PatternFill("solid", "DDEBF7"), Font(bold=True), align="left")
    for j, l in enumerate(layers, 2):
        val = metrics_results[l][m_name]
        c_val = ws_dash.cell(i, j, val)
        fmt = "0.00%" if any(x in m_name for x in ["CAGR", "Return", "Drawdown", "Win", "Volatility", "Gain", "Loss", "VaR", "CVaR"]) else "0.00"
        c_val.number_format = fmt
        style_cell(c_val, font=Font(color="006100" if val > 0 else "9C0006" if val < 0 else "000000"))

# 2. Churning Analysis
ws_churn = wb.create_sheet("Churning_Analysis")
c_hdrs = ["Month", "Ex_Ante_Sharpe", "Port_Beta", "Stock_Count"] + [f"{l} Add" for l in ['Base','ST','EMA','COMBO','ULTRA']] + [f"{l} Rem" for l in ['Base','ST','EMA','COMBO','ULTRA']]
for j, h in enumerate(c_hdrs, 1): style_cell(ws_churn.cell(1, j, h), HDR_FILL, WHITE_FONT)
for i, row in df_churn.iterrows():
    ws_churn.cell(i+2, 1, row['Month'])
    ws_churn.cell(i+2, 2, row['Ex_Ante_Sharpe']).number_format = "0.00"
    ws_churn.cell(i+2, 3, row['Port_Beta']).number_format = "0.00"
    ws_churn.cell(i+2, 4, row['Stock_Count']).number_format = "0"
    for j, k in enumerate(c_hdrs[4:], 5): ws_churn.cell(i+2, j, row[k])

# Enhanced Churning Statistics
r_off = len(df_churn) + 4
ws_churn.cell(r_off, 1, "Churning Statistics Summary").font = Font(bold=True, size=12)
ws_churn.cell(r_off+1, 1, "Layer"); ws_churn.cell(r_off+1, 2, "Min Add"); ws_churn.cell(r_off+1, 3, "Max Add"); ws_churn.cell(r_off+1, 4, "Avg Add")
for i, l in enumerate(['Base','ST','EMA','COMBO','ULTRA'], 1):
    ws_churn.cell(r_off+1+i, 1, l)
    ws_churn.cell(r_off+1+i, 2, df_churn[f"{l} Add"].min())
    ws_churn.cell(r_off+1+i, 3, df_churn[f"{l} Add"].max())
    ws_churn.cell(r_off+1+i, 4, df_churn[f"{l} Add"].mean()).number_format = "0.0"

# 2.5 Detailed Monthly Summary
ws_mon = wb.create_sheet("Detailed_Monthly_Summary")
m_hdrs = list(df_sum.columns)
ws_mon.cell(1, 1, "Month"); style_cell(ws_mon.cell(1, 1), HDR_FILL, WHITE_FONT)
for j, h in enumerate(m_hdrs, 2): style_cell(ws_mon.cell(1, j, h), HDR_FILL, WHITE_FONT)
for i, (idx, row) in enumerate(df_sum.iterrows(), 2):
    ws_mon.cell(i, 1, idx)
    for j, col in enumerate(m_hdrs, 2):
        c = ws_mon.cell(i, j, row[col])
        if any(x in col for x in ["Count", "Added", "Removed"]): c.number_format = "0"
        elif any(x in col for x in ["Beta", "Sharpe"]): c.number_format = "0.00"
        else: c.number_format = "0.00%"

# 3. Monthly Heatmap
ws_heat = wb.create_sheet("Monthly_Heatmap")
df_sum_heat = pd.DataFrame(monthly_summary)
df_sum_heat['Date'] = pd.to_datetime(df_sum_heat['Month'])
df_sum_heat.set_index('Date', inplace=True)
df_sum_heat['Year'] = df_sum_heat.index.year
df_sum_heat['Month_Num'] = df_sum_heat.index.month
heatmap_data = df_sum_heat.pivot_table(index='Year', columns='Month_Num', values='ULTRA_HEDGE')
months_short = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

for j, m in enumerate(months_short, 2): style_cell(ws_heat.cell(1, j, m), HDR_FILL, WHITE_FONT)
for i, year in enumerate(heatmap_data.index, 2):
    style_cell(ws_heat.cell(i, 1, year), HDR_FILL, WHITE_FONT)
    for j, month in enumerate(range(1, 13), 2):
        val = heatmap_data.loc[year, month] if month in heatmap_data.columns else None
        if val is not None:
            c = ws_heat.cell(i, j, val); c.number_format = "0.0%"
            color = "C6EFCE" if val > 0.02 else "FFEB9C" if val > -0.02 else "FFC7CE"
            style_cell(c, PatternFill("solid", color))

# 3. Strategy Definitions & Math Methodology
ws_def = wb.create_sheet("Strategy_Definitions")
ws_def.column_dimensions['A'].width = 30
ws_def.column_dimensions['B'].width = 80

def_hdr = PatternFill("solid", "203764") # Navy
def_sub = PatternFill("solid", "D9E1F2") # Light blue

defs = [
    ("Metric/Layer", "Logic Implementation: What we did in this layer"),
    ("Base SIM", "Selected stocks using Treynor-Black optimization. Calculated monthly/daily Betas, Residual Variance (Sigma2), and Excess Return to Beta (ERB). Applied cut-off (Ci) and capped individual weights at 10% for institutional diversification."),
    ("ST Filter (Momentum)", "Filtered the Base portfolio by applying the Supertrend (10,3) indicator on monthly historical data. We only retained stocks showing a Bullish signal (Trend = 1) to eliminate downward momentum."),
    ("EMA Filter (Trend)", "Filtered the Base portfolio using a 40-period Exponential Moving Average (EMA). We implemented a rule to only hold stocks trading above their EMA-40, ensuring alignment with the long-term primary trend."),
    ("COMBO Filter", "Implemented a dual-validation logic: stocks must pass BOTH the Supertrend Bullishness AND the EMA-40 Trend requirement. This 'Double-Lock' system significantly filters out market noise and weak rallies."),
    ("ULTRA Layer", "A concentrated high-conviction momentum layer. Currently synchronized with the COMBO selection, it serves as the base for aggressive defense mechanisms when market breadth deteriorates."),
    ("COMBO + Hedge", "Integrated an active NIFTY Futures hedge. We monitor Multi-Timeframe (15m, 30m, 1h, 1d) Supertrend alignment. When all 4 timeframes turn bearish, we short Nifty Futures proportional to the Portfolio Beta to neutralize market risk."),
    ("ULTRA Defense", "Implemented an aggressive 'Multiplier' hedge. If market breadth (Combo stocks / Base stocks) drops below 50%, we apply a 2.0x Beta multiplier to the Nifty short position, providing deep protection during system-wide crashes."),
    ("Alpha (Institutional)", "The excess return generated over the benchmark (CAGR_p - CAGR_bm)."),
    ("XIRR", "Extended Internal Rate of Return, accounting for the time value of money and capital flows."),
    ("Calmar Ratio", "CAGR / |Max Drawdown|. Measures return per unit of drawdown risk."),
    ("Sortino Ratio", "(CAGR - RF) / Downside Deviation. Penalizes only negative volatility."),
    ("Expectancy", "The average amount you expect to win (or lose) per trade: (Win% * AvgGain) + (Loss% * AvgLoss)."),
    ("VaR 99%", "95%/99% confidence that losses will not exceed this value in a single month."),
    ("CVaR (Expected Shortfall)", "The average loss in the extreme tail (beyond VaR) cases."),
    ("Portfolio Beta", "Weighted average of individual stock betas: Σ (Weight_i * Beta_i). Measures sensitivity to Nifty 50."),
    ("Ex-Ante Sharpe", "Forward-looking Sharpe ratio derived from the expected returns and the daily covariance matrix of selected stocks."),
    ("Survival Filter", "Institutional look-ahead check ensuring stock data availability in the trade month, effectively removing delisting and liquidity risks.")
]

for i, (m, d) in enumerate(defs, 1):
    ws_def.cell(i, 1, m); ws_def.cell(i, 2, d)
    style_cell(ws_def.cell(i, 1), fill=def_hdr if i==1 else def_sub, font=WHITE_FONT if i==1 else Font(bold=True))
    style_cell(ws_def.cell(i, 2), fill=def_hdr if i==1 else None, font=WHITE_FONT if i==1 else None, align="left")
    ws_def.cell(i, 2).alignment = Alignment(wrap_text=True)

# 4. Detailed Monthly Portfolio Sheets
df_audit = pd.DataFrame(portfolio_audit_log)
STATUS_COLORS = {'Added': "C6EFCE", 'Remained': "DDEBF7"} # Green for Added, Light Blue for Remained

for t_month in df_audit['Trade_Month'].unique():
    # A month is 'LIVE' if it is the latest Trade_Month and we are currently in or before that month
    curr_m = datetime.now().strftime('%Y-%m')
    is_live = (t_month >= curr_m)
    sheet_name = f"Port_{t_month}" if not is_live else f"LIVE_PERF_{t_month}"
    ws_port = wb.create_sheet(sheet_name)
    m_data = df_audit[df_audit['Trade_Month'] == t_month].sort_values('Weight', ascending=False)
    
    # Summary Info at Top
    sum_row = df_sum.loc[t_month]
    ws_port.merge_cells('A1:M1')
    hdr_text = f"PORTFOLIO FOR {t_month} (Based on {m_data['Month'].iloc[0]} Data)"
    if is_live: hdr_text = f"*** LIVE PERFORMANCE {t_month} *** (Monitoring Active Portfolio)"
    ws_port['A1'] = f"{hdr_text}  |  Beta: {sum_row['Port_Beta']:.3f}  |  Sharpe: {sum_row['Ex_Ante_Sharpe']:.3f}  |  Stocks: {int(sum_row['Stock_Count'])}"
    style_cell(ws_port['A1'], def_hdr, WHITE_FONT)
    
    p_hdrs = ["Stock", "SIM Weight", "ST Wt", "EMA Wt", "COMBO Wt", "ULTRA Wt", "Beta", "ERB", "Status", "Qty", "Prev_Qty", "Delta", "Action"]
    p_cols = ["Symbol", "Weight", "ST_Wt", "EMA_Wt", "COMBO_Wt", "ULTRA_Wt", "Beta", "ERB", "Status", "Qty", "Prev_Qty", "Delta_Qty", "Action"]
    
    for j, h in enumerate(p_hdrs, 1): 
        style_cell(ws_port.cell(2, j, h), PatternFill("solid", "ACB9CA"), Font(bold=True))
        ws_port.column_dimensions[gcl(j)].width = 12
    ws_port.column_dimensions['M'].width = 18 # Action column wider
    
    for idx, row in m_data.iterrows():
        r_idx = ws_port.max_row + 1
        for j, col in enumerate(p_cols, 1):
            c = ws_port.cell(r_idx, j, row[col])
            if "Wt" in col or col == "Weight": c.number_format = "0.00%"
            elif col in ["Beta", "ERB"]: c.number_format = "0.00"
            elif col == "Action" and "BUY" in str(row[col]): style_cell(c, fill=PatternFill("solid", "C6EFCE"), font=Font(bold=True))
            elif col == "Action" and "SELL" in str(row[col]): style_cell(c, fill=PatternFill("solid", "FFC7CE"), font=Font(bold=True))
            
            # Special Styling for Status
            if col == "Status":
                fill = PatternFill("solid", STATUS_COLORS.get(row['Status'], "FFFFFF"))
                style_cell(c, fill=fill)
            elif col not in ["Action"]:
                style_cell(c)

# ==============================================================================
# [Sub-Excel] DEEP-DIVE INSTITUTIONAL REPORT
# ==============================================================================

print(f"[*] Generating Deep-Dive Report: {DEEP_DIVE_FILE}")
wb_dd = Workbook()

# 1. Rolling Stats & Visual Data
ws_ts = wb_dd.active
ws_ts.title = "Time_Series_Analytics"
ts_hdrs = ["Month", "Equity_Base", "Equity_Combo_H", "Equity_Ultra_H", "Drawdown_Base", "Drawdown_Combo_H", "Bench_Equity"]
for j, h in enumerate(ts_hdrs, 1): style_cell(ws_ts.cell(1, j, h), HDR_FILL, WHITE_FONT)

eq_base, eq_ch, eq_uh, eq_bench = 1.0, 1.0, 1.0, 1.0
eq_series_base, eq_series_ch, eq_series_uh, eq_series_bench = [1.0], [1.0], [1.0], [1.0]

for i, (idx, row) in enumerate(df_sum.iterrows(), 2):
    eq_base *= (1 + row['Base'])
    eq_ch *= (1 + row['COMBO_HEDGE'])
    eq_uh *= (1 + row['ULTRA_HEDGE'])
    eq_bench *= (1 + row['Bench'])
    eq_series_base.append(eq_base)
    eq_series_ch.append(eq_ch)
    eq_series_uh.append(eq_uh)
    
    ws_ts.cell(i, 1, idx)
    ws_ts.cell(i, 2, eq_base).number_format = "0.00"
    ws_ts.cell(i, 3, eq_ch).number_format = "0.00"
    ws_ts.cell(i, 4, eq_uh).number_format = "0.00"
    
    # Calculate Drawdown
    dd_base = (eq_base / max(eq_series_base)) - 1
    ws_ts.cell(i, 5, dd_base).number_format = "0.00%"
    ws_ts.cell(i, 6, (eq_ch / max(eq_series_ch)) - 1).number_format = "0.00%"
    ws_ts.cell(i, 7, eq_bench).number_format = "0.00"

# 2. Layer-wise Monthly Heatmaps
# Includes all filters (ST, EMA) and all Hedged Defense layers
for layer in ['Base', 'ST', 'EMA', 'COMBO', 'ULTRA', 'COMBO_HEDGE', 'ULTRA_HEDGE']:
    ws_l_heat = wb_dd.create_sheet(f"Heatmap_{layer}")
    df_l = pd.DataFrame(monthly_summary)
    df_l['Date'] = pd.to_datetime(df_l['Month'])
    df_l.set_index('Date', inplace=True)
    l_heat = df_l.pivot_table(index=df_l.index.year, columns=df_l.index.month, values=layer)
    
    for j, m in enumerate(months_short, 2): style_cell(ws_l_heat.cell(1, j, m), HDR_FILL, WHITE_FONT)
    for i, year in enumerate(l_heat.index, 2):
        style_cell(ws_l_heat.cell(i, 1, year), HDR_FILL, WHITE_FONT)
        for j, month in enumerate(range(1, 13), 2):
            val = l_heat.loc[year, month] if month in l_heat.columns else None
            if val is not None:
                c = ws_l_heat.cell(i, j, val); c.number_format = "0.0%"
                color = "C6EFCE" if val > 0.02 else "FFEB9C" if val > -0.02 else "FFC7CE"
                style_cell(c, PatternFill("solid", color))

# 3. Full Execution Log
ws_log = wb_dd.create_sheet("Execution_History")
log_hdrs = ["Month", "Symbol", "Action", "Qty", "Price", "Return"]
for j, h in enumerate(log_hdrs, 1): style_cell(ws_log.cell(1, j, h), HDR_FILL, WHITE_FONT)
for i, row in df_audit.iterrows():
    if "HOLD" not in str(row['Action']):
        r_idx = ws_log.max_row + 1
        ws_log.cell(r_idx, 1, row['Month'])
        ws_log.cell(r_idx, 2, row['Symbol'])
        ws_log.cell(r_idx, 3, row['Action'])
        ws_log.cell(r_idx, 4, row['Qty'])
        ws_log.cell(r_idx, 5, row['Buy_Px']).number_format = "0.00"
        ws_log.cell(r_idx, 6, row['Return']).number_format = "0.00%"

print(f"[v] Reports Saved: {OUTPUT_FILE} and {DEEP_DIVE_FILE}")
wb.save(OUTPUT_FILE)
wb_dd.save(DEEP_DIVE_FILE)
